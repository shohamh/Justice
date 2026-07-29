from __future__ import annotations

import json

import openpyxl
import pytest
from sqlalchemy import delete, select

from app.db.models import BugReport, DutyManagerScope, SystemSetting
from app.routes.config_export import _write_bug_reports, _write_system_settings
from app.services.hierarchy import create_node
from tests.helpers import auth_headers, create_soldier


def _make_duty_manager(session, personal_number: str):
    """A duty_manager actor needs at least one DutyManagerScope row —
    require_duty_manager_or_admin checks for scope, not just role."""
    dm = create_soldier(session, personal_number=personal_number, role="duty_manager")
    node = create_node(session, level="team", name=f"node_{personal_number}", parent_id=None)
    session.add(DutyManagerScope(duty_manager_id=dm.id, hierarchy_node_id=node.id))
    session.commit()
    return dm


@pytest.fixture(autouse=True)
def _clear_system_settings_for_export_tests(admin_session) -> None:
    """Clear system_settings before each test so the export tests work correctly."""
    admin_session.execute(delete(SystemSetting))
    admin_session.commit()


def _rows(ws):
    return [
        [c.value for c in row]
        for row in ws.iter_rows(min_row=2)
        if any(c.value is not None for c in row)
    ]


def test_write_system_settings_writes_key_and_json_value(admin_session):
    admin_session.add(SystemSetting(key="algorithm.max_duties_per_window", value=8, updated_by=None))
    admin_session.add(SystemSetting(key="telegram.enabled", value=True, updated_by=None))
    admin_session.commit()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_system_settings(wb, admin_session)

    rows = {r[0]: r[1] for r in _rows(wb["system_settings"])}
    assert rows["algorithm.max_duties_per_window"] == "8"
    assert rows["telegram.enabled"] == "true"


def test_write_system_settings_excludes_hidden_keys(admin_session):
    admin_session.add(SystemSetting(key="system.holding_node_id", value="abc", updated_by=None))
    admin_session.commit()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_system_settings(wb, admin_session)

    assert _rows(wb["system_settings"]) == []


def test_write_bug_reports_resolves_reporter_and_serializes_json_columns(admin_session):
    reporter = create_soldier(admin_session, personal_number="5556667", role="soldier")
    admin_session.add(BugReport(
        reporter_id=reporter.id, description="בעיה", severity="high", route="/x", status="open",
        nav_history=[{"path": "/a"}], audit_snapshot=None, user_snapshot={"role": "soldier"},
    ))
    admin_session.commit()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_bug_reports(wb, admin_session)

    row = _rows(wb["bug_reports"])[0]
    header = [c.value for c in wb["bug_reports"][1]]
    data = dict(zip(header, row))
    assert data["reporter_personal_number"] == "5556667"
    assert data["description"] == "בעיה"
    assert json.loads(data["nav_history_json"]) == [{"path": "/a"}]
    assert data["audit_snapshot_json"] == ""
    assert json.loads(data["user_snapshot_json"]) == {"role": "soldier"}


def test_export_by_duty_manager_drops_system_settings_and_bug_reports(client, admin_session):
    """Critical Fix 1: system_settings/bug_reports are admin-only end to
    end. A duty-manager's export (no sheets param → defaults to all six)
    silently drops these two rather than erroring the whole export."""
    import io as _io

    dm = _make_duty_manager(admin_session, "dm_export_1")

    resp = client.get("/api/config/export", headers=auth_headers(dm))
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    assert "system_settings" not in wb.sheetnames
    assert "bug_reports" not in wb.sheetnames
    assert "duty_types" in wb.sheetnames


def test_export_by_duty_manager_explicit_request_still_drops_admin_only_sheets(client, admin_session):
    """A request for *only* admin-only sheets by a non-admin has nothing left
    to export once they're dropped — openpyxl can't save a workbook with zero
    visible sheets, so this is a 400 rather than a 200 with an empty file."""
    dm = _make_duty_manager(admin_session, "dm_export_2")

    resp = client.get(
        "/api/config/export?sheets=system_settings,bug_reports", headers=auth_headers(dm)
    )
    assert resp.status_code == 400
    assert resp.json()["detail"] == "no_exportable_sheets"


def test_export_by_admin_includes_system_settings_and_bug_reports(client, admin_session):
    import io as _io

    admin = create_soldier(admin_session, personal_number="admin_export_1", role="admin")
    admin_session.commit()

    resp = client.get("/api/config/export", headers=auth_headers(admin))
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    assert "system_settings" in wb.sheetnames
    assert "bug_reports" in wb.sheetnames
