from __future__ import annotations

import io
import logging
import secrets
import uuid
from datetime import date as date_type
from typing import Any, Literal

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.audit.writer import write_audit
from app.auth.deps import require_duty_manager_or_admin, require_password_changed
from app.db.models import (
    DutyAssignment,
    DutyLocation,
    DutyShift,
    DutyShiftNodeQuota,
    DutyType,
    HierarchyNode,
    NotificationType,
    ShiftTemplate,
    Soldier,
)
from app.db.session import get_session
from app.services.import_parsers._shared_parsing import parse_bool as _parse_bool
from app.services.import_parsers._shared_parsing import parse_date as _parse_date
from app.services.import_scope import is_node_in_actor_scope
from app.services.notifications import create_notification

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/import", tags=["import"])


# ── Row models ─────────────────────────────────────────────────────────────────

class SoldierRowPreview(BaseModel):
    row: int
    action: Literal["new", "update", "error"]
    personal_number: str
    full_name: str
    rank: str | None
    gender: str | None
    is_officer: bool | None
    hierarchy_node_id: uuid.UUID | None
    hierarchy_node_name: str | None
    enrolled_at: str | None
    enlistment_date: str | None
    phone: str | None
    email: str | None
    existing_id: uuid.UUID | None
    errors: list[str]


class AssignmentRowPreview(BaseModel):
    row: int
    action: Literal["new", "error"]
    personal_number: str
    duty_type_name: str
    start_date: str
    end_date: str
    is_reserve: bool
    resolved_soldier_id: uuid.UUID | None
    resolved_duty_type_id: uuid.UUID | None
    errors: list[str]


class PreviewResult(BaseModel):
    soldiers: list[SoldierRowPreview]
    assignments: list[AssignmentRowPreview]


# ── Apply models ───────────────────────────────────────────────────────────────

class ApplySoldierRow(BaseModel):
    row: int
    action: Literal["new", "update", "skip"]
    personal_number: str
    full_name: str
    rank: str | None
    gender: str | None
    is_officer: bool | None
    hierarchy_node_id: uuid.UUID | None
    enrolled_at: str | None
    enlistment_date: str | None
    phone: str | None
    email: str | None
    existing_id: uuid.UUID | None


class ApplyAssignmentRow(BaseModel):
    row: int
    action: Literal["new", "skip"]
    resolved_soldier_id: uuid.UUID
    resolved_duty_type_id: uuid.UUID
    start_date: str
    end_date: str
    is_reserve: bool


class ApplyRequest(BaseModel):
    soldiers: list[ApplySoldierRow]
    assignments: list[ApplyAssignmentRow]


class ApplyResult(BaseModel):
    created: int
    updated: int
    skipped: int
    errors: list[str]


# ── Preview endpoint ───────────────────────────────────────────────────────────

@router.post("/preview", response_model=PreviewResult)
async def preview(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    actor: Soldier = Depends(require_duty_manager_or_admin),
):
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="file_too_large")
    # Validate XLSX magic bytes (PK signature for ZIP format)
    if content[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="invalid_file_type")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid_xlsx")

    soldiers_by_pn: dict[str, Soldier] = {
        s.personal_number: s
        for s in session.execute(select(Soldier)).scalars().all()
    }
    nodes_by_name: dict[str, HierarchyNode] = {
        n.name: n
        for n in session.execute(select(HierarchyNode)).scalars().all()
    }
    duty_types_by_name: dict[str, DutyType] = {
        dt.name: dt
        for dt in session.execute(select(DutyType)).scalars().all()
    }

    soldiers = _parse_soldiers_sheet(wb, soldiers_by_pn, nodes_by_name)
    assignments = _parse_assignments_sheet(wb, soldiers_by_pn, duty_types_by_name)

    return PreviewResult(soldiers=soldiers, assignments=assignments)


def _parse_soldiers_sheet(wb, soldiers_by_pn, nodes_by_name) -> list[SoldierRowPreview]:
    if "soldiers" not in wb.sheetnames:
        return []
    ws = wb["soldiers"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    results = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        data = dict(zip(headers, row))
        errors: list[str] = []
        pn = str(data.get("personal_number") or "").strip()
        full_name = str(data.get("full_name") or "").strip()
        if not pn:
            errors.append("personal_number is required")
        if not full_name:
            errors.append("full_name is required")

        node_name = str(data.get("hierarchy_node_name") or "").strip() or None
        node = nodes_by_name.get(node_name) if node_name else None
        if node_name and not node:
            errors.append(f"hierarchy_node_name '{node_name}' not found")

        existing = soldiers_by_pn.get(pn)
        action: Literal["new", "update", "error"] = "error" if errors else ("update" if existing else "new")

        results.append(SoldierRowPreview(
            row=i,
            action=action,
            personal_number=pn,
            full_name=full_name,
            rank=str(data.get("rank") or "").strip() or None,
            gender=str(data.get("gender") or "").strip() or None,
            is_officer=_parse_bool(data.get("is_officer")),
            hierarchy_node_id=node.id if node else None,
            hierarchy_node_name=node_name,
            enrolled_at=_parse_date(data.get("enrolled_at")),
            enlistment_date=_parse_date(data.get("enlistment_date")),
            phone=str(data.get("phone") or "").strip() or None,
            email=str(data.get("email") or "").strip() or None,
            existing_id=existing.id if existing else None,
            errors=errors,
        ))
    return results


def _parse_assignments_sheet(wb, soldiers_by_pn, duty_types_by_name) -> list[AssignmentRowPreview]:
    if "assignments" not in wb.sheetnames:
        return []
    ws = wb["assignments"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    results = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        data = dict(zip(headers, row))
        errors: list[str] = []
        pn = str(data.get("personal_number") or "").strip()
        dt_name = str(data.get("duty_type_name") or "").strip()
        start = _parse_date(data.get("start_date"))
        end = _parse_date(data.get("end_date"))

        soldier = soldiers_by_pn.get(pn)
        if not pn or not soldier:
            errors.append(f"personal_number '{pn}' not found")
        dt = duty_types_by_name.get(dt_name)
        if not dt_name or not dt:
            errors.append(f"duty_type_name '{dt_name}' not found")
        if not start:
            errors.append("start_date is required (dd.mm.yyyy)")
        if not end:
            errors.append("end_date is required (dd.mm.yyyy)")

        results.append(AssignmentRowPreview(
            row=i,
            action="error" if errors else "new",
            personal_number=pn,
            duty_type_name=dt_name,
            start_date=start or "",
            end_date=end or "",
            is_reserve=_parse_bool(data.get("is_reserve")) or False,
            resolved_soldier_id=soldier.id if soldier else None,
            resolved_duty_type_id=dt.id if dt else None,
            errors=errors,
        ))
    return results


# ── Apply endpoint ─────────────────────────────────────────────────────────────

def _out_of_scope_rows(session: Session, actor: Soldier, req: ApplyRequest) -> list[str]:
    """Return a list of human-readable row descriptions the actor may not import,
    because they touch a hierarchy node outside the actor's DutyManagerScope.
    Empty for admins."""
    if actor.role == "admin":
        return []
    errors: list[str] = []
    for row in req.soldiers:
        if row.action == "skip":
            continue
        if row.action == "new":
            if not is_node_in_actor_scope(session=session, actor=actor, node_id=row.hierarchy_node_id):
                errors.append(f"soldier row {row.row}: hierarchy node out of your scope")
        elif row.action == "update" and row.existing_id:
            existing = session.get(Soldier, row.existing_id)
            current_node_id = existing.hierarchy_node_id if existing else None
            if not is_node_in_actor_scope(session=session, actor=actor, node_id=current_node_id):
                errors.append(f"soldier row {row.row}: soldier's current node is out of your scope")
            if row.hierarchy_node_id is not None and not is_node_in_actor_scope(
                session=session, actor=actor, node_id=row.hierarchy_node_id
            ):
                errors.append(f"soldier row {row.row}: destination node out of your scope")
    for row in req.assignments:
        if row.action == "skip":
            continue
        soldier = session.get(Soldier, row.resolved_soldier_id)
        node_id = soldier.hierarchy_node_id if soldier else None
        if not is_node_in_actor_scope(session=session, actor=actor, node_id=node_id):
            errors.append(f"assignment row {row.row}: soldier out of your scope")
    return errors


@router.post("/apply", response_model=ApplyResult)
def apply(
    req: ApplyRequest,
    session: Session = Depends(get_session),
    actor: Soldier = Depends(require_duty_manager_or_admin),
):
    out_of_scope = _out_of_scope_rows(session, actor, req)
    if out_of_scope:
        raise HTTPException(status_code=403, detail={"out_of_scope_rows": out_of_scope})

    from app.auth.password import hash_password

    created = updated = skipped = 0
    errors: list[str] = []
    created_assignments: list[DutyAssignment] = []

    try:
        for row in req.soldiers:
            if row.action == "skip":
                skipped += 1
                continue
            if row.action == "new":
                new_soldier = Soldier(
                    personal_number=row.personal_number,
                    full_name=row.full_name,
                    password_hash=hash_password(secrets.token_hex(16)),
                    must_change_password=True,
                    rank=row.rank,
                    gender=row.gender,
                    is_officer=row.is_officer,
                    hierarchy_node_id=row.hierarchy_node_id,
                    phone=row.phone,
                    email=row.email,
                )
                if row.enrolled_at:
                    new_soldier.enrolled_at = date_type.fromisoformat(row.enrolled_at)
                if row.enlistment_date:
                    new_soldier.enlistment_date = date_type.fromisoformat(row.enlistment_date)
                session.add(new_soldier)
                created += 1
            elif row.action == "update" and row.existing_id:
                s = session.get(Soldier, row.existing_id)
                if s:
                    s.full_name = row.full_name
                    if row.rank is not None:
                        s.rank = row.rank
                    if row.gender is not None:
                        s.gender = row.gender
                    if row.is_officer is not None:
                        s.is_officer = row.is_officer
                    if row.hierarchy_node_id is not None:
                        s.hierarchy_node_id = row.hierarchy_node_id
                    if row.phone is not None:
                        s.phone = row.phone
                    if row.email is not None:
                        s.email = row.email
                    if row.enrolled_at:
                        s.enrolled_at = date_type.fromisoformat(row.enrolled_at)
                    if row.enlistment_date:
                        s.enlistment_date = date_type.fromisoformat(row.enlistment_date)
                    updated += 1

        session.flush()

        # Assignments
        for row in req.assignments:
            if row.action == "skip":
                skipped += 1
                continue
            loc = session.execute(select(DutyLocation).limit(1)).scalar_one_or_none()
            if loc is None:
                errors.append(f"Row {row.row}: no duty location exists — cannot import assignment")
                continue
            assignment = DutyAssignment(
                soldier_id=row.resolved_soldier_id,
                duty_type_id=row.resolved_duty_type_id,
                duty_location_id=loc.id,
                start_date=date_type.fromisoformat(row.start_date),
                end_date=date_type.fromisoformat(row.end_date),
                status="published",
                is_reserve=row.is_reserve,
            )
            session.add(assignment)
            session.flush()
            created_assignments.append(assignment)
            created += 1

        write_audit(
            session, actor_id=actor.id, action="import.excel_apply", entity_type="import_batch",
            after={"created": created, "updated": updated, "skipped": skipped,
                   "created_assignment_ids": [str(a.id) for a in created_assignments]},
        )
        session.commit()
    except Exception as exc:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(exc))

    # The import itself (soldiers/assignments/audit row) already committed
    # above. Each notification is sent — and committed — independently so
    # that one bad row (e.g. row 2 of 10) doesn't prevent the rest of the
    # batch from being attempted: a failure here must never turn a
    # successful import into an unhandled 500 for the client, nor silently
    # drop notifications for unrelated assignments.
    for a in created_assignments:
        try:
            create_notification(
                session, soldier_id=a.soldier_id,
                type=NotificationType.assignment_created,
                title="שיבוץ חדש נוצר עבורך (ייבוא Excel)",
                reference_type="duty_assignment", reference_id=a.id,
                actor_id=actor.id,
            )
            session.commit()
        except Exception:
            session.rollback()
            logger.error(
                "Failed to send assignment_created notification for assignment %s "
                "after Excel import apply",
                a.id,
                exc_info=True,
            )

    return ApplyResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ── Template download ──────────────────────────────────────────────────────────

@router.get("/template")
def download_template():
    """Download an example workbook for the active import pipeline.

    Matches the `v1_standard` parser's expected sheets — see
    app/services/import_parsers/v1_standard.py.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_s = wb.create_sheet("soldiers")
    ws_s.append(["personal_number", "full_name", "rank", "gender", "is_officer",
                  "hierarchy_node_name", "enrolled_at", "enlistment_date", "phone", "email"])
    ws_s.append(["12345", "ישראל ישראלי", "רב\"ט", "m", "false", "מדור א", "01.01.2022", "01.03.2020", "050-1234567", "israel@example.com"])
    ws_s.append(["23456", "משה כהן", "סמל", "m", "false", "מדור א", "15.02.2022", "10.05.2020", "050-2345678", ""])
    ws_s.append(["34567", "דנה לוי", "רס\"ל", "f", "false", "מדור ב", "01.03.2022", "20.06.2020", "050-3456789", "dana@example.com"])
    ws_s.append(["45678", "יעל אברהם", "סרן", "f", "true", "מדור ב", "10.01.2021", "01.09.2018", "050-4567890", "yael@example.com"])
    ws_s.append(["56789", "אבי מזרחי", "רב\"ט", "m", "false", "מדור א", "05.04.2022", "12.07.2020", "", ""])
    ws_s.append(["67890", "נועה שרון", "טוראי", "f", "false", "מדור ב", "20.05.2023", "01.01.2023", "050-6789012", ""])
    ws_s.append(["78901", "רון פרידמן", "רב\"ט", "m", "false", "מדור א", "15.06.2022", "01.09.2020", "050-7890123", ""])
    ws_s.append(["89012", "עידן ברק", "סגן", "m", "true", "מדור ב", "01.02.2021", "01.03.2019", "050-8901234", "idan@example.com"])

    ws_d = wb.create_sheet("duty_shifts")
    ws_d.append(["duty_type_name", "duty_location_name", "start_date", "end_date",
                  "start_time", "end_time", "required_count", "node_quotas", "notes"])
    ws_d.append(["שמירה", "שער ראשי", "15.06.2024", "16.06.2024", "20:00", "06:00", 4, "מדור א:2;מדור ב:2", "הצטיידות במקלע"])
    ws_d.append(["שמירה", "שער ראשי", "16.06.2024", "17.06.2024", "20:00", "06:00", 4, "מדור א:2;מדור ב:2", ""])
    ws_d.append(["מטבח", "מטבח מרכזי", "15.06.2024", "16.06.2024", "05:00", "13:00", 2, "מדור ב:2", ""])
    ws_d.append(["סיור", "היקף מחנה", "17.06.2024", "18.06.2024", "22:00", "04:00", 3, "מדור א:1;מדור ב:2", "נדרש רכב"])
    ws_d.append(["משמר לילה", "מגדל שמירה צפוני", "18.06.2024", "19.06.2024", "23:00", "05:00", 2, "", "תורנות רגישה — לוודא תדריך"])
    ws_d.append(["תורנות סוף שבוע", "שער ראשי", "21.06.2024", "23.06.2024", "", "", 6, "מדור א:3;מדור ב:3", ""])

    ws_a = wb.create_sheet("assignments")
    ws_a.append(["personal_number", "full_name", "duty_type_name", "duty_location_name",
                  "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes"])
    ws_a.append(["12345", "ישראל ישראלי", "שמירה", "שער ראשי", "15.06.2024", "16.06.2024", "20:00", "06:00", "false", ""])
    ws_a.append(["23456", "משה כהן", "שמירה", "שער ראשי", "15.06.2024", "16.06.2024", "20:00", "06:00", "true", "מחליף תורן"])

    ws_loc = wb.create_sheet("duty_locations")
    ws_loc.append(["name", "base", "active"])
    ws_loc.append(["שער ראשי", "בסיס א", "true"])
    ws_loc.append(["מטבח מרכזי", "בסיס א", "true"])

    ws_h = wb.create_sheet("hierarchy")
    ws_h.append(["name", "level", "parent_name", "commander_personal_number", "commander_name", "duty_managers"])
    ws_h.append(["אוגדה 1", "division", "", "", "", ""])
    ws_h.append(["מדור א", "group", "אוגדה 1", "12345", "ישראל ישראלי", "23456:משה כהן"])

    ws_dt = wb.create_sheet("duty_types")
    ws_dt.append([
        "name", "score_per_day", "description", "active", "reserve_ratio", "reserve_minimum",
        "is_external", "contact_name", "contact_phone", "start_time", "end_time",
        "instructions", "eligible_units", "requirements_json",
    ])
    ws_dt.append([
        "שמירה", "1.50", "שמירה בשער הראשי", "true", "0.200", "2",
        "false", "דני", "050-1234567", "20:00", "06:00",
        "הצטיידות במקלע", "מדור א", "{}",
    ])

    ws_et = wb.create_sheet("exemption_types")
    ws_et.append(["name", "description", "is_global", "is_medical", "is_commander_exemption", "applies_to_duty_types"])
    ws_et.append(["פטור רפואי", "אישור רופא", "false", "true", "false", "שמירה"])

    ws_tpl = wb.create_sheet("shift_templates")
    ws_tpl.append([
        "name", "duty_type_name", "duty_location_name", "recurrence_type", "weekdays",
        "start_time", "end_time", "required_count", "auto_roll", "auto_roll_until",
        "duration_days", "notes", "eligible_units",
    ])
    ws_tpl.append([
        "שמירה לילה", "שמירה", "שער ראשי", "weekly", "1,3",
        "20:00", "06:00", 2, "false", "", 1, "", "מדור א",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="import_template.xlsx"'},
    )


# ── Export current data ─────────────────────────────────────────────────────────

EXPORT_DATA_SHEETS = ["soldiers", "duty_shifts", "assignments", "shift_templates"]


@router.get("/export")
def export_current_data(
    sheets: str | None = None,
    session: Session = Depends(get_session),
    actor: Soldier = Depends(require_duty_manager_or_admin),
):
    """Dump current soldiers/duty_shifts/assignments/shift_templates into the
    same layout as the import template, for a full export -> edit -> re-import
    round trip. Assignments with no linked `duty_shift_id` (not tied to a
    shift instance) are omitted — they have no composite key to export.

    `sheets` is an optional comma-separated subset of EXPORT_DATA_SHEETS;
    defaults to all four when omitted, matching /config/export's convention.
    """
    requested = (
        {s.strip() for s in sheets.split(",")} if sheets else set(EXPORT_DATA_SHEETS)
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    nodes_by_id = {n.id: n for n in session.execute(select(HierarchyNode)).scalars()}
    duty_types_by_id = {dt.id: dt for dt in session.execute(select(DutyType)).scalars()}
    locations_by_id = {loc.id: loc for loc in session.execute(select(DutyLocation)).scalars()}

    if "soldiers" in requested:
        ws_s = wb.create_sheet("soldiers")
        ws_s.append(["personal_number", "full_name", "rank", "gender", "is_officer",
                      "hierarchy_node_name", "enrolled_at", "enlistment_date", "phone", "email"])
        for s in session.execute(select(Soldier)).scalars():
            node = nodes_by_id.get(s.hierarchy_node_id) if s.hierarchy_node_id else None
            ws_s.append([
                s.personal_number, s.full_name, s.rank, s.gender,
                "" if s.is_officer is None else ("true" if s.is_officer else "false"),
                node.name if node else "",
                s.enrolled_at.strftime("%d.%m.%Y") if s.enrolled_at else "",
                s.enlistment_date.strftime("%d.%m.%Y") if s.enlistment_date else "",
                s.phone or "", s.email or "",
            ])

    # `assignments` needs shift lookups even when the `duty_shifts` sheet itself isn't requested.
    if "duty_shifts" in requested or "assignments" in requested:
        shifts = session.execute(select(DutyShift)).scalars().all()
    else:
        shifts = []

    if "duty_shifts" in requested:
        quotas_by_shift: dict[uuid.UUID, list[str]] = {}
        for quota, node_name in session.execute(
            select(DutyShiftNodeQuota, HierarchyNode.name).join(
                HierarchyNode, DutyShiftNodeQuota.hierarchy_node_id == HierarchyNode.id
            )
        ):
            quotas_by_shift.setdefault(quota.duty_shift_id, []).append(f"{node_name}:{quota.count}")

        ws_d = wb.create_sheet("duty_shifts")
        ws_d.append(["duty_type_name", "duty_location_name", "start_date", "end_date",
                      "start_time", "end_time", "required_count", "node_quotas", "notes"])
        for shift in shifts:
            dt = duty_types_by_id.get(shift.duty_type_id)
            loc = locations_by_id.get(shift.duty_location_id)
            ws_d.append([
                dt.name if dt else "", loc.name if loc else "",
                shift.start_date.strftime("%d.%m.%Y"), shift.end_date.strftime("%d.%m.%Y"),
                shift.start_time, shift.end_time, shift.required_count,
                ";".join(quotas_by_shift.get(shift.id, [])), shift.notes or "",
            ])

    if "assignments" in requested:
        shifts_by_id = {shift.id: shift for shift in shifts}
        soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
        ws_a = wb.create_sheet("assignments")
        ws_a.append(["personal_number", "full_name", "duty_type_name", "duty_location_name",
                      "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes"])
        for a in session.execute(select(DutyAssignment)).scalars():
            if a.duty_shift_id is None:
                continue
            shift = shifts_by_id.get(a.duty_shift_id)
            if shift is None:
                continue
            soldier = soldiers_by_id.get(a.soldier_id)
            dt = duty_types_by_id.get(shift.duty_type_id)
            loc = locations_by_id.get(shift.duty_location_id)
            ws_a.append([
                soldier.personal_number if soldier else "",
                soldier.full_name if soldier else "",
                dt.name if dt else "", loc.name if loc else "",
                shift.start_date.strftime("%d.%m.%Y"), shift.end_date.strftime("%d.%m.%Y"),
                shift.start_time, shift.end_time,
                "true" if a.is_reserve else "false",
                a.notes or "",
            ])

    if "shift_templates" in requested:
        ws_tpl = wb.create_sheet("shift_templates")
        ws_tpl.append([
            "name", "duty_type_name", "duty_location_name", "recurrence_type", "weekdays",
            "start_time", "end_time", "required_count", "auto_roll", "auto_roll_until",
            "duration_days", "notes", "eligible_units",
        ])
        for tpl in session.execute(select(ShiftTemplate)).scalars():
            dt = duty_types_by_id.get(tpl.duty_type_id)
            loc = locations_by_id.get(tpl.duty_location_id)
            eligible = ", ".join(
                nodes_by_id[nid].name for nid in (tpl.eligible_node_ids or []) if nid in nodes_by_id
            )
            ws_tpl.append([
                tpl.name,
                dt.name if dt else "",
                loc.name if loc else "",
                tpl.recurrence_type,
                ",".join(str(d) for d in tpl.weekdays),
                tpl.start_time, tpl.end_time, tpl.required_count,
                "true" if tpl.auto_roll else "false",
                tpl.auto_roll_until.strftime("%d.%m.%Y") if tpl.auto_roll_until else "",
                tpl.duration_days,
                tpl.notes or "",
                eligible,
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="export.xlsx"'},
    )
