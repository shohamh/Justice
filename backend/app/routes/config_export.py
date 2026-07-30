from __future__ import annotations

import io
import json

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth.deps import require_duty_manager_or_admin
from app.db.models import (
    BugReport,
    DutyLocation,
    DutyManagerScope,
    DutyType,
    ExemptionDutyTypeMap,
    ExemptionType,
    HierarchyNode,
    Soldier,
    SystemSetting,
)
from app.db.session import get_session
from app.services.settings_loader import _HIDDEN_KEYS

router = APIRouter(prefix="/config", tags=["config-export"])

ALL_SHEETS = ["duty_types", "duty_locations", "hierarchy", "exemption_types", "system_settings", "bug_reports"]


def _write_duty_locations(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("duty_locations")
    ws.append(["name", "base", "active"])
    for loc in session.execute(select(DutyLocation)).scalars():
        ws.append([loc.name, loc.base, loc.active])


def _write_hierarchy(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("hierarchy")
    ws.append(["name", "level", "parent_name", "commander_personal_number", "commander_name", "duty_managers"])
    nodes = list(session.execute(select(HierarchyNode)).scalars())
    nodes_by_id = {n.id: n for n in nodes}
    soldier_ids = {n.commander_id for n in nodes if n.commander_id}
    dm_rows = list(session.execute(select(DutyManagerScope)).scalars())
    soldier_ids |= {r.duty_manager_id for r in dm_rows}
    soldiers_by_id = {
        s.id: s for s in session.execute(select(Soldier).where(Soldier.id.in_(soldier_ids))).scalars()
    } if soldier_ids else {}
    dm_by_node: dict = {}
    for r in dm_rows:
        dm_by_node.setdefault(r.hierarchy_node_id, []).append(r.duty_manager_id)

    for n in nodes:
        parent_name = nodes_by_id[n.parent_id].name if n.parent_id in nodes_by_id else ""
        commander = soldiers_by_id.get(n.commander_id) if n.commander_id else None
        dm_cell = ";".join(
            f"{soldiers_by_id[sid].personal_number}:{soldiers_by_id[sid].full_name}"
            for sid in dm_by_node.get(n.id, []) if sid in soldiers_by_id
        )
        ws.append([
            n.name, n.level, parent_name,
            commander.personal_number if commander else "",
            commander.full_name if commander else "",
            dm_cell,
        ])


def _write_duty_types(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("duty_types")
    ws.append([
        "name", "score_per_day", "description", "active", "reserve_ratio", "reserve_minimum",
        "is_external", "contact_name", "contact_phone", "start_time", "end_time",
        "instructions", "eligible_units", "requirements_json",
    ])
    nodes_by_id = {n.id: n for n in session.execute(select(HierarchyNode)).scalars()}
    for dt in session.execute(select(DutyType)).scalars():
        eligible = ", ".join(
            nodes_by_id[nid].name for nid in (dt.eligible_node_ids or []) if nid in nodes_by_id
        )
        ws.append([
            dt.name, str(dt.score_per_day), dt.description, dt.active,
            str(dt.reserve_ratio), dt.reserve_minimum, dt.is_external,
            dt.contact_name, dt.contact_phone,
            dt.start_time.strftime("%H:%M") if dt.start_time else "",
            dt.end_time.strftime("%H:%M") if dt.end_time else "",
            dt.instructions, eligible,
            json.dumps(dt.requirements, ensure_ascii=False) if dt.requirements else "",
        ])


def _write_exemption_types(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("exemption_types")
    ws.append(["name", "description", "is_global", "is_medical", "is_commander_exemption", "applies_to_duty_types"])
    duty_types_by_id = {dt.id: dt for dt in session.execute(select(DutyType)).scalars()}
    map_rows = list(session.execute(select(ExemptionDutyTypeMap)).scalars())
    applies_by_et: dict = {}
    for m in map_rows:
        applies_by_et.setdefault(m.exemption_type_id, []).append(m.duty_type_id)

    for et in session.execute(select(ExemptionType)).scalars():
        applies = ", ".join(
            duty_types_by_id[dtid].name
            for dtid in applies_by_et.get(et.id, []) if dtid in duty_types_by_id
        )
        ws.append([
            et.name, et.description, et.is_global, et.is_medical, et.is_commander_exemption, applies,
        ])


def _write_system_settings(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("system_settings")
    ws.append(["key", "value_json"])
    for setting in session.execute(select(SystemSetting)).scalars():
        if setting.key in _HIDDEN_KEYS:
            continue
        ws.append([setting.key, json.dumps(setting.value, ensure_ascii=False)])


def _write_bug_reports(wb: openpyxl.Workbook, session: Session) -> None:
    ws = wb.create_sheet("bug_reports")
    ws.append([
        "id", "reporter_personal_number", "description", "severity", "route", "status",
        "created_at", "nav_history_json", "audit_snapshot_json", "user_snapshot_json",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    for br in session.execute(select(BugReport)).scalars():
        reporter = soldiers_by_id.get(br.reporter_id)
        ws.append([
            str(br.id),
            reporter.personal_number if reporter else "",
            br.description, br.severity, br.route, br.status,
            br.created_at.isoformat() if br.created_at else "",
            json.dumps(br.nav_history, ensure_ascii=False) if br.nav_history else "",
            json.dumps(br.audit_snapshot, ensure_ascii=False) if br.audit_snapshot else "",
            json.dumps(br.user_snapshot, ensure_ascii=False) if br.user_snapshot else "",
        ])


_WRITERS = {
    "duty_locations": _write_duty_locations,
    "hierarchy": _write_hierarchy,
    "duty_types": _write_duty_types,
    "exemption_types": _write_exemption_types,
    "system_settings": _write_system_settings,
    "bug_reports": _write_bug_reports,
}


@router.get("/export")
def export_config(
    sheets: str | None = None,
    session: Session = Depends(get_session),
    actor: Soldier = Depends(require_duty_manager_or_admin),
):
    requested = [s.strip() for s in sheets.split(",")] if sheets else ALL_SHEETS
    requested = [s for s in requested if s in _WRITERS]
    if actor.role != "admin":
        # system_settings/bug_reports are admin-only end to end: a duty manager
        # can still export the other sheets they're allowed to see, but these
        # two are silently dropped rather than erroring the whole export.
        requested = [s for s in requested if s not in ("system_settings", "bug_reports")]

    if not requested:
        # openpyxl can't save a workbook with zero visible sheets, and an
        # export with nothing in it isn't a meaningful response either —
        # this only happens when every requested sheet was explicitly
        # admin-only and the actor isn't an admin.
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="no_exportable_sheets")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in requested:
        _WRITERS[sheet_name](wb, session)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="config_export.xlsx"'},
    )
