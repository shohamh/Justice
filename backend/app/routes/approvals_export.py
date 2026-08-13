from __future__ import annotations

import io

import openpyxl
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth.authz import can_see_private
from app.auth.deps import require_duty_manager_or_admin
from app.db.models import (
    ExemptionRequest, ExemptionRequestFile, ExemptionType, HierarchyNode,
    PersonalConstraint, RangeAssignment, RangeExcusalRequest, RangeEvent, RangeLocation,
    Soldier, SoldierEnrollmentRequest, SoldierExemption,
    SoldierFieldUpdate, SoldierRangeQualification, SwapCandidate, SwapManagerApproval, SwapRequest,
)
from app.db.session import get_session

router = APIRouter(prefix="/approvals", tags=["approvals-export"])

ALL_SHEETS = [
    "swap_requests", "exemption_requests", "soldier_field_updates",
    "soldier_enrollment_requests", "personal_constraints", "soldier_exemptions",
    "soldier_range_qualifications", "range_excusal_requests",
]


def _soldier_label(soldiers_by_id: dict, soldier_id) -> tuple[str, str]:
    s = soldiers_by_id.get(soldier_id)
    return (s.personal_number, s.full_name) if s else ("", "")


def _write_personal_constraints(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("personal_constraints")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "start_date", "end_date",
        "reason", "status", "decided_by_personal_number", "decision_note", "created_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    visibility_cache: dict = {}
    for c in session.execute(select(PersonalConstraint)).scalars():
        pn, name = _soldier_label(soldiers_by_id, c.soldier_id)
        decided_pn = _soldier_label(soldiers_by_id, c.decided_by)[0] if c.decided_by else ""
        soldier = soldiers_by_id.get(c.soldier_id)
        if soldier is None:
            include_reason = False
        elif c.soldier_id in visibility_cache:
            include_reason = visibility_cache[c.soldier_id]
        else:
            include_reason = can_see_private(session, actor, soldier)
            visibility_cache[c.soldier_id] = include_reason
        reason = c.reason if include_reason else None
        ws.append([
            str(c.id), pn, name, c.start_date.isoformat(), c.end_date.isoformat(),
            reason, c.status, decided_pn, c.decision_note, c.created_at.isoformat(),
        ])


def _write_soldier_field_updates(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("soldier_field_updates")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "field_name", "new_value",
        "previous_value", "status", "decided_by_personal_number", "decision_note", "created_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    for u in session.execute(select(SoldierFieldUpdate)).scalars():
        pn, name = _soldier_label(soldiers_by_id, u.soldier_id)
        decided_pn = _soldier_label(soldiers_by_id, u.decided_by)[0] if u.decided_by else ""
        ws.append([
            str(u.id), pn, name, u.field_name, u.new_value, u.previous_value,
            u.status, decided_pn, u.decision_note, u.created_at.isoformat(),
        ])


def _write_soldier_enrollment_requests(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("soldier_enrollment_requests")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "requested_node_name",
        "status", "decided_by_personal_number", "decision_note", "created_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    nodes_by_id = {n.id: n for n in session.execute(select(HierarchyNode)).scalars()}
    for r in session.execute(select(SoldierEnrollmentRequest)).scalars():
        pn, name = _soldier_label(soldiers_by_id, r.soldier_id)
        decided_pn = _soldier_label(soldiers_by_id, r.decided_by)[0] if r.decided_by else ""
        node = nodes_by_id.get(r.requested_node_id)
        ws.append([
            str(r.id), pn, name, node.name if node else "",
            r.status, decided_pn, r.decision_note, r.created_at.isoformat(),
        ])


def _write_soldier_exemptions(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("soldier_exemptions")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "exemption_type_name",
        "start_date", "end_date", "reason", "granted_by_personal_number", "granted_at",
        "revoked_at", "revoked_by_personal_number", "revoke_reason",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    exemption_types_by_id = {et.id: et for et in session.execute(select(ExemptionType)).scalars()}
    for e in session.execute(select(SoldierExemption)).scalars():
        pn, name = _soldier_label(soldiers_by_id, e.soldier_id)
        et = exemption_types_by_id.get(e.exemption_type_id)
        granted_pn = _soldier_label(soldiers_by_id, e.granted_by)[0] if e.granted_by else ""
        revoked_pn = _soldier_label(soldiers_by_id, e.revoked_by)[0] if e.revoked_by else ""
        ws.append([
            str(e.id), pn, name, et.name if et else "",
            e.start_date.isoformat(), e.end_date.isoformat() if e.end_date else "",
            e.reason, granted_pn, e.granted_at.isoformat(),
            e.revoked_at.isoformat() if e.revoked_at else "", revoked_pn, e.revoke_reason,
        ])


def _write_exemption_requests(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("exemption_requests")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "exemption_type_name",
        "start_date", "end_date", "reason", "status",
        "commander_approved_by_personal_number", "decided_by_personal_number",
        "decision_note", "files", "created_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    exemption_types_by_id = {et.id: et for et in session.execute(select(ExemptionType)).scalars()}
    files_by_request: dict = {}
    for f in session.execute(select(ExemptionRequestFile)).scalars():
        files_by_request.setdefault(f.exemption_request_id, []).append(f.file_name)
    visibility_cache: dict = {}
    for r in session.execute(select(ExemptionRequest)).scalars():
        pn, name = _soldier_label(soldiers_by_id, r.soldier_id)
        et = exemption_types_by_id.get(r.exemption_type_id)
        commander_pn = _soldier_label(soldiers_by_id, r.commander_approved_by)[0] if r.commander_approved_by else ""
        decided_pn = _soldier_label(soldiers_by_id, r.decided_by)[0] if r.decided_by else ""
        files = ", ".join(files_by_request.get(r.id, []))
        soldier = soldiers_by_id.get(r.soldier_id)
        if soldier is None:
            include_sensitive = False
        elif r.soldier_id in visibility_cache:
            include_sensitive = visibility_cache[r.soldier_id]
        else:
            include_sensitive = can_see_private(session, actor, soldier)
            visibility_cache[r.soldier_id] = include_sensitive
        reason = r.reason if include_sensitive else None
        ws.append([
            str(r.id), pn, name, et.name if et else "",
            r.start_date.isoformat() if r.start_date else "", r.end_date.isoformat() if r.end_date else "",
            reason, r.status, commander_pn, decided_pn, r.decision_note, files, r.created_at.isoformat(),
        ])


def _write_swap_requests(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    """One row per SwapRequest x SwapCandidate (or one bare row if the request
    has no candidates yet), since a request can now have several candidates
    instead of exactly one target/covering soldier. `target_personal_number`
    holds a still-unanswered invited candidate (mirrors the old pre-response
    target_soldier_id); `covering_personal_number` holds a candidate who has
    accepted, applied, or self-claimed from the marketplace (mirrors the old
    covering_soldier_id). `covering_side_approved` is that candidate's own
    `soldier_side_approved`. `approval_log` is the union of the shared
    requester-side decisions and this candidate's own covering-side decisions,
    so each row is a self-contained record of that candidate's approval state."""
    ws = wb.create_sheet("swap_requests")
    ws.append([
        "id", "requesting_personal_number", "requesting_name", "target_personal_number",
        "covering_personal_number", "duty_date", "status", "reason",
        "requester_side_approved", "covering_side_approved",
        "rejected_by_personal_number", "decision_note", "approval_log", "created_at", "updated_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}

    # Keyed by (swap_request_id, swap_candidate_id) — candidate_id is None for
    # the shared requester-side rows, and set for a specific candidate's
    # covering-side rows, matching SwapManagerApproval.swap_candidate_id.
    decisions_by_key: dict[tuple, list[str]] = {}
    for d in session.execute(select(SwapManagerApproval)).scalars():
        if not (d.approved or d.rejected):
            continue
        person = soldiers_by_id.get(d.commander_id)
        if person is None:
            continue
        outcome = "approved" if d.approved else "rejected"
        at = d.approved_at if d.approved else d.rejected_at
        decisions_by_key.setdefault((d.swap_request_id, d.swap_candidate_id), []).append(
            f"{d.side}:{d.approver_kind}:{person.personal_number}:{outcome}:{at.isoformat() if at else ''}"
        )

    candidates_by_request: dict = {}
    for c in session.execute(select(SwapCandidate)).scalars():
        candidates_by_request.setdefault(c.swap_request_id, []).append(c)

    for r in session.execute(select(SwapRequest)).scalars():
        requesting_pn, requesting_name = _soldier_label(soldiers_by_id, r.requesting_soldier_id)
        rejected_pn = _soldier_label(soldiers_by_id, r.rejected_by)[0] if r.rejected_by else ""
        requester_log = decisions_by_key.get((r.id, None), [])
        candidates = sorted(candidates_by_request.get(r.id, []), key=lambda c: c.created_at)

        if not candidates:
            ws.append([
                str(r.id), requesting_pn, requesting_name, "", "",
                r.duty_date.isoformat(), r.status, r.reason,
                r.requester_side_approved, None,
                rejected_pn, r.decision_note, ";".join(requester_log),
                r.created_at.isoformat(), r.updated_at.isoformat(),
            ])
            continue

        for c in candidates:
            pn, _name = _soldier_label(soldiers_by_id, c.soldier_id)
            still_unanswered_invite = c.source == "invited" and c.status == "pending"
            target_pn = pn if still_unanswered_invite else ""
            covering_pn = "" if still_unanswered_invite else pn
            approval_log = ";".join(requester_log + decisions_by_key.get((r.id, c.id), []))
            ws.append([
                str(r.id), requesting_pn, requesting_name, target_pn, covering_pn,
                r.duty_date.isoformat(), r.status, r.reason,
                r.requester_side_approved, c.soldier_side_approved,
                rejected_pn, r.decision_note, approval_log,
                r.created_at.isoformat(), r.updated_at.isoformat(),
            ])


def _write_soldier_range_qualifications(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("soldier_range_qualifications")
    ws.append(["id", "soldier_personal_number", "soldier_name", "range_type", "valid_until"])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    for q in session.execute(select(SoldierRangeQualification)).scalars():
        pn, name = _soldier_label(soldiers_by_id, q.soldier_id)
        ws.append([str(q.id), pn, name, q.range_type.value, q.valid_until.isoformat()])


def _write_range_excusal_requests(wb: openpyxl.Workbook, session: Session, actor: Soldier) -> None:
    ws = wb.create_sheet("range_excusal_requests")
    ws.append([
        "id", "soldier_personal_number", "soldier_name", "requested_by_personal_number",
        "hierarchy_node_name", "range_type", "date", "range_location_name",
        "reason", "status", "decided_by_personal_number", "decision_note", "requested_at",
    ])
    soldiers_by_id = {s.id: s for s in session.execute(select(Soldier)).scalars()}
    nodes_by_id = {n.id: n for n in session.execute(select(HierarchyNode)).scalars()}
    locations_by_id = {loc.id: loc for loc in session.execute(select(RangeLocation)).scalars()}
    events_by_id = {ev.id: ev for ev in session.execute(select(RangeEvent)).scalars()}
    assignments_by_id = {a.id: a for a in session.execute(select(RangeAssignment)).scalars()}
    for r in session.execute(select(RangeExcusalRequest)).scalars():
        assignment = assignments_by_id.get(r.range_assignment_id) if r.range_assignment_id else None
        soldier_id = assignment.soldier_id if assignment else None
        pn, _name = _soldier_label(soldiers_by_id, soldier_id) if soldier_id else ("", "")
        requested_pn = _soldier_label(soldiers_by_id, r.requested_by)[0] if r.requested_by else ""
        decided_pn = _soldier_label(soldiers_by_id, r.decided_by)[0] if r.decided_by else ""
        event = events_by_id.get(r.range_event_id) if r.range_event_id else None
        node = nodes_by_id.get(event.hierarchy_node_id) if event else None
        loc = locations_by_id.get(event.range_location_id) if event else None
        ws.append([
            str(r.id), pn, "", requested_pn,
            node.name if node else "", event.range_type.value if event else "",
            event.date.isoformat() if event else "", loc.name if loc else "",
            r.reason, r.status.value, decided_pn, r.decision_note, r.requested_at.isoformat(),
        ])


_WRITERS = {
    "swap_requests": _write_swap_requests,
    "exemption_requests": _write_exemption_requests,
    "soldier_field_updates": _write_soldier_field_updates,
    "soldier_enrollment_requests": _write_soldier_enrollment_requests,
    "personal_constraints": _write_personal_constraints,
    "soldier_exemptions": _write_soldier_exemptions,
    "soldier_range_qualifications": _write_soldier_range_qualifications,
    "range_excusal_requests": _write_range_excusal_requests,
}


@router.get("/export")
def export_approvals(
    sheets: str | None = None,
    session: Session = Depends(get_session),
    actor: Soldier = Depends(require_duty_manager_or_admin),
):
    requested = [s.strip() for s in sheets.split(",")] if sheets else ALL_SHEETS
    requested = [s for s in requested if s in _WRITERS]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in requested:
        _WRITERS[sheet_name](wb, session, actor)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="approvals_export.xlsx"'},
    )
