from __future__ import annotations

from typing import Any

import openpyxl

from app.services.excel_bilingual import canonical_headers, canonical_sheet_name, resolve_sheet_name
from app.services.import_parsers._shared_parsing import parse_bool as _parse_bool
from app.services.import_parsers._shared_parsing import parse_date as _parse_date
from app.services.import_parsers.registry import register
from app.services.import_parsers.schema import (
    ImportAssignmentRow,
    ImportBugReportRow,
    ImportDutyLocationRow,
    ImportDutyShiftRow,
    ImportDutyTypeRow,
    ImportExemptionRequestRow,
    ImportExemptionTypeRow,
    ImportHierarchyNodeRow,
    ImportNodeQuota,
    ImportPersonalConstraintRow,
    ImportRangeAssignmentRow,
    ImportRangeEventRow,
    ImportRangeExcusalRequestRow,
    ImportRangeLocationRow,
    ImportRankAdvancementIntervalRow,
    ImportShiftTemplateRow,
    ImportSoldierEnrollmentRequestRow,
    ImportSoldierExemptionRow,
    ImportSoldierFieldUpdateRow,
    ImportSoldierRangeQualificationRow,
    ImportSoldierRow,
    ImportSwapRequestRow,
    ImportSystemSettingRow,
    ParsedImportData,
)

KNOWN_SHEETS = {
    "soldiers", "duty_shifts", "assignments", "duty_locations", "hierarchy",
    "duty_types", "exemption_types", "shift_templates",
    "swap_requests", "exemption_requests", "soldier_field_updates",
    "soldier_enrollment_requests", "personal_constraints", "soldier_exemptions",
    "system_settings", "bug_reports",
    "range_locations", "range_events", "range_assignments",
    "soldier_range_qualifications", "range_excusal_requests",
    "rank_advancement_intervals",
}


def _sheet_rows(wb: openpyxl.Workbook, name: str) -> list[dict[str, Any]]:
    """Read a sheet's rows as dicts keyed by lowercased header, skipping blank rows.

    Ported convention from app/routes/import_excel.py's per-sheet parsers:
    header row lowercased, data starts at row 2, all-None rows are skipped.
    """
    resolved = resolve_sheet_name(wb.sheetnames, name)
    if resolved is None:
        return []
    ws = wb[resolved]
    raw_headers = [
        str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    headers = canonical_headers(name, raw_headers)
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        out.append({"_row": i, **dict(zip(headers, row))})
    return out


def _parse_name_list(raw: Any) -> list[str]:
    """Parse a comma-separated list of names, stripping whitespace.

    Used for applies_to_duty_type_names, eligible_unit_names, etc.
    Empty cell or whitespace-only cell returns empty list.
    """
    s = str(raw or "").strip()
    if not s:
        return []
    return [name.strip() for name in s.split(",") if name.strip()]


def _parse_int_list(raw: Any) -> list[int]:
    """Parse a comma-separated list of integers (used for `weekdays`).

    Non-integer entries are skipped silently — malformed weekday values are
    caught by the resolver's recurrence_type/weekdays validation, not here.
    """
    s = str(raw or "").strip()
    if not s:
        return []
    out = []
    for part in s.split(","):
        part = part.strip()
        if part.isdigit():
            out.append(int(part))
    return out


def _parse_node_quotas(raw: Any, source_row: int) -> tuple[list[ImportNodeQuota], list[str]]:
    """Parse the new `node_quotas` column: "node_name:count;node_name:count".

    Malformed entries (missing colon, or a non-integer count) are skipped
    individually rather than crashing the whole import or vanishing silently;
    each produces a row-tagged warning string.
    """
    s = str(raw or "").strip()
    if not s:
        return [], []
    quotas: list[ImportNodeQuota] = []
    warnings: list[str] = []
    for part in s.split(";"):
        part = part.strip()
        if not part:
            continue
        if ":" not in part:
            warnings.append(
                f"שורה {source_row}: ערך מכסה שגוי '{part}' — הפורמט הנדרש הוא 'שם_יחידה:כמות'"
            )
            continue
        name, count_s = part.rsplit(":", 1)
        try:
            count = int(count_s.strip())
        except ValueError:
            warnings.append(
                f"שורה {source_row}: ערך מכסה שגוי '{part}' — הפורמט הנדרש הוא 'שם_יחידה:כמות'"
            )
            continue
        quotas.append(ImportNodeQuota(node_name=name.strip(), count=count))
    return quotas, warnings


def _parse_duty_manager_refs(raw: Any, source_row: int) -> tuple[list[str], list[str]]:
    """Parse `personal_number:full_name;personal_number:full_name` into a list
    of raw `"pn:name"` strings (resolved later against real soldiers) — same
    `;`-then-`:` convention as `_parse_node_quotas`. Malformed entries (missing
    colon) produce a row-tagged warning and are skipped individually."""
    s = str(raw or "").strip()
    if not s:
        return [], []
    refs: list[str] = []
    warnings: list[str] = []
    for part in s.split(";"):
        part = part.strip()
        if not part:
            continue
        if ":" not in part:
            warnings.append(
                f"שורה {source_row}: ערך אחראי תורנות שגוי '{part}' — הפורמט הנדרש הוא 'מספר_אישי:שם_מלא'"
            )
            continue
        refs.append(part)
    return refs, warnings


class V1StandardParser:
    """Standard v1 layout: `soldiers`, `duty_shifts`, `assignments`,
    `duty_locations`, `hierarchy`, `duty_types`, `exemption_types`,
    `shift_templates`.
    """

    id = "v1_standard"
    label = "תבנית סטנדרטית (v1)"

    def detect(self, wb: openpyxl.Workbook) -> float:
        canonical = {canonical_sheet_name(name) for name in wb.sheetnames}
        matches = KNOWN_SHEETS & canonical
        if not matches:
            return 0.0
        return min(1.0, 0.5 + 0.2 * len(matches))

    def parse(self, wb: openpyxl.Workbook) -> ParsedImportData:
        warnings: list[str] = []

        soldiers = [
            ImportSoldierRow(
                source_row=r["_row"],
                personal_number=str(r.get("personal_number") or "").strip(),
                full_name=str(r.get("full_name") or "").strip(),
                rank=str(r.get("rank") or "").strip() or None,
                gender=str(r.get("gender") or "").strip() or None,
                is_officer=_parse_bool(r.get("is_officer")),
                hierarchy_node_name=str(r.get("hierarchy_node_name") or "").strip() or None,
                enrolled_at=_parse_date(r.get("enrolled_at")),
                enlistment_date=_parse_date(r.get("enlistment_date")),
                unit_join_date=_parse_date(r.get("unit_join_date")),
                phone=str(r.get("phone") or "").strip() or None,
                email=str(r.get("email") or "").strip() or None,
                food_type=str(r.get("food_type") or "").strip() or None,
                food_constraints=str(r.get("food_constraints") or "").strip() or None,
                profile_picture_url=str(r.get("profile_picture_url") or "").strip() or None,
                telegram_chat_id=(
                    int(r["telegram_chat_id"])
                    if r.get("telegram_chat_id") not in (None, "")
                    else None
                ),
                telegram_username=str(r.get("telegram_username") or "").strip() or None,
                telegram_is_verified=_parse_bool(r.get("telegram_is_verified")),
                telegram_notifications_enabled=_parse_bool(r.get("telegram_notifications_enabled")),
                telegram_verified_at=str(r.get("telegram_verified_at") or "").strip() or None,
                is_career=_parse_bool(r.get("is_career")),
                next_rank_date=_parse_date(r.get("next_rank_date")),
                next_rank_date_overridden=_parse_bool(r.get("next_rank_date_overridden")),
                current_rank_since=_parse_date(r.get("current_rank_since")),
                bahad1_graduate=_parse_bool(r.get("bahad1_graduate")),
                rank_track=str(r.get("rank_track") or "").strip() or None,
                has_military_driving_license=_parse_bool(r.get("has_military_driving_license")),
                military_driving_license_expiry=_parse_date(r.get("military_driving_license_expiry")),
                mandatory_end_date=_parse_date(r.get("mandatory_end_date")),
                discharge_date=_parse_date(r.get("discharge_date")),
                last_mitvahim_date=_parse_date(r.get("last_mitvahim_date")),
                last_alal_date=_parse_date(r.get("last_alal_date")),
                left_at=_parse_date(r.get("left_at")),
            )
            for r in _sheet_rows(wb, "soldiers")
        ]

        duty_shifts = []
        for r in _sheet_rows(wb, "duty_shifts"):
            node_quotas, node_quota_warnings = _parse_node_quotas(r.get("node_quotas"), r["_row"])
            warnings.extend(node_quota_warnings)
            duty_shifts.append(
                ImportDutyShiftRow(
                    source_row=r["_row"],
                    duty_type_name=str(r.get("duty_type_name") or "").strip(),
                    duty_location_name=str(r.get("duty_location_name") or "").strip(),
                    start_date=_parse_date(r.get("start_date")) or "",
                    end_date=_parse_date(r.get("end_date")) or "",
                    start_time=str(r.get("start_time") or "").strip() or None,
                    end_time=str(r.get("end_time") or "").strip() or None,
                    required_count=int(r.get("required_count") or 1),
                    reserve_count_override=int(r["reserve_count_override"]) if r.get("reserve_count_override") else None,
                    node_quotas=node_quotas,
                    notes=str(r.get("notes") or "").strip() or None,
                )
            )

        assignments = [
            ImportAssignmentRow(
                source_row=r["_row"],
                personal_number=str(r.get("personal_number") or "").strip(),
                full_name=str(r.get("full_name") or "").strip(),
                duty_type_name=str(r.get("duty_type_name") or "").strip(),
                duty_location_name=str(r.get("duty_location_name") or "").strip(),
                start_date=_parse_date(r.get("start_date")) or "",
                end_date=_parse_date(r.get("end_date")) or "",
                start_time=str(r.get("start_time") or "").strip() or None,
                end_time=str(r.get("end_time") or "").strip() or None,
                is_reserve=_parse_bool(r.get("is_reserve")) or False,
                notes=str(r.get("notes") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "assignments")
        ]

        duty_locations = [
            ImportDutyLocationRow(
                source_row=r["_row"],
                name=str(r.get("name") or "").strip(),
                base=str(r.get("base") or "").strip() or None,
                active=_parse_bool(r.get("active")),
            )
            for r in _sheet_rows(wb, "duty_locations")
        ]

        exemption_types = [
            ImportExemptionTypeRow(
                source_row=r["_row"],
                name=str(r.get("name") or "").strip(),
                description=str(r.get("description") or "").strip() or None,
                is_global=_parse_bool(r.get("is_global")),
                is_medical=_parse_bool(r.get("is_medical")),
                is_commander_exemption=_parse_bool(r.get("is_commander_exemption")),
                active=_parse_bool(r.get("active")),
                applies_to_duty_type_names=_parse_name_list(r.get("applies_to_duty_types")),
                forbids_weapons=_parse_bool(r.get("forbids_weapons")),
            )
            for r in _sheet_rows(wb, "exemption_types")
        ]

        hierarchy = []
        for r in _sheet_rows(wb, "hierarchy"):
            dm_refs, dm_warnings = _parse_duty_manager_refs(r.get("duty_managers"), r["_row"])
            warnings.extend(dm_warnings)
            hierarchy.append(
                ImportHierarchyNodeRow(
                    source_row=r["_row"],
                    name=str(r.get("name") or "").strip(),
                    level=str(r.get("level") or "").strip(),
                    parent_name=str(r.get("parent_name") or "").strip() or None,
                    commander_personal_number=str(r.get("commander_personal_number") or "").strip() or None,
                    commander_name=str(r.get("commander_name") or "").strip() or None,
                    duty_manager_refs=dm_refs,
                )
            )
        duty_types = [
            ImportDutyTypeRow(
                source_row=r["_row"],
                name=str(r.get("name") or "").strip(),
                score_per_day=str(r.get("score_per_day") or "").strip(),
                description=str(r.get("description") or "").strip() or None,
                active=_parse_bool(r.get("active")),
                reserve_ratio=str(r.get("reserve_ratio") or "").strip() or None,
                reserve_minimum=(
                    int(r["reserve_minimum"])
                    if r.get("reserve_minimum") is not None and str(r["reserve_minimum"]).strip() != ""
                    else None
                ),
                is_external=_parse_bool(r.get("is_external")),
                contact_name=str(r.get("contact_name") or "").strip() or None,
                contact_phone=str(r.get("contact_phone") or "").strip() or None,
                start_time=str(r.get("start_time") or "").strip() or None,
                end_time=str(r.get("end_time") or "").strip() or None,
                instructions=str(r.get("instructions") or "").strip() or None,
                eligible_unit_names=_parse_name_list(r.get("eligible_units")),
                requires_weapon=_parse_bool(r.get("requires_weapon")),
                required_range_type=str(r.get("required_range_type") or "").strip() or None,
                requirements_json=str(r.get("requirements_json") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "duty_types")
        ]

        shift_templates = [
            ImportShiftTemplateRow(
                source_row=r["_row"],
                name=str(r.get("name") or "").strip(),
                duty_type_name=str(r.get("duty_type_name") or "").strip(),
                duty_location_name=str(r.get("duty_location_name") or "").strip(),
                recurrence_type=str(r.get("recurrence_type") or "").strip() or None,
                weekdays=_parse_int_list(r.get("weekdays")),
                start_time=str(r.get("start_time") or "").strip() or None,
                end_time=str(r.get("end_time") or "").strip() or None,
                required_count=(
                    int(r["required_count"])
                    if r.get("required_count") is not None and str(r["required_count"]).strip() != ""
                    else None
                ),
                # `or None` first: an empty-string cell (as opposed to a genuinely
                # unset/None cell) would otherwise parse_bool to False rather than
                # "blank" — normalize both to None so a blank auto_roll cell means
                # "leave unchanged" on an update row, matching required_count/
                # duration_days/recurrence_type above.
                auto_roll=_parse_bool(r.get("auto_roll") or None),
                auto_roll_until=_parse_date(r.get("auto_roll_until")),
                duration_days=(
                    int(r["duration_days"])
                    if r.get("duration_days") is not None and str(r["duration_days"]).strip() != ""
                    else None
                ),
                notes=str(r.get("notes") or "").strip() or None,
                eligible_unit_names=_parse_name_list(r.get("eligible_units")),
            )
            for r in _sheet_rows(wb, "shift_templates")
        ]

        swap_requests = [
            ImportSwapRequestRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                requesting_personal_number=str(r.get("requesting_personal_number") or "").strip(),
                target_personal_number=str(r.get("target_personal_number") or "").strip() or None,
                covering_personal_number=str(r.get("covering_personal_number") or "").strip() or None,
                duty_date=_parse_date(r.get("duty_date")) or "",
                status=str(r.get("status") or "").strip(),
                reason=str(r.get("reason") or "").strip() or None,
                requester_side_approved=_parse_bool(r.get("requester_side_approved")),
                covering_side_approved=_parse_bool(r.get("covering_side_approved")),
                rejected_by_personal_number=str(r.get("rejected_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
                approval_log=str(r.get("approval_log") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "swap_requests")
        ]

        exemption_requests = [
            ImportExemptionRequestRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                exemption_type_name=str(r.get("exemption_type_name") or "").strip(),
                start_date=_parse_date(r.get("start_date")) or "",
                end_date=_parse_date(r.get("end_date")),
                reason=str(r.get("reason") or "").strip() or None,
                status=str(r.get("status") or "").strip(),
                commander_approved_by_personal_number=str(r.get("commander_approved_by_personal_number") or "").strip() or None,
                decided_by_personal_number=str(r.get("decided_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
                files=str(r.get("files") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "exemption_requests")
        ]

        soldier_field_updates = [
            ImportSoldierFieldUpdateRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                field_name=str(r.get("field_name") or "").strip(),
                new_value=str(r.get("new_value") or "").strip(),
                previous_value=str(r.get("previous_value") or "").strip() or None,
                status=str(r.get("status") or "").strip(),
                decided_by_personal_number=str(r.get("decided_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "soldier_field_updates")
        ]

        soldier_enrollment_requests = [
            ImportSoldierEnrollmentRequestRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                requested_node_name=str(r.get("requested_node_name") or "").strip(),
                status=str(r.get("status") or "").strip(),
                decided_by_personal_number=str(r.get("decided_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "soldier_enrollment_requests")
        ]

        personal_constraints = [
            ImportPersonalConstraintRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                start_date=_parse_date(r.get("start_date")) or "",
                end_date=_parse_date(r.get("end_date")) or "",
                reason=str(r.get("reason") or "").strip() or None,
                status=str(r.get("status") or "").strip(),
                decided_by_personal_number=str(r.get("decided_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "personal_constraints")
        ]

        soldier_exemptions = [
            ImportSoldierExemptionRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                exemption_type_name=str(r.get("exemption_type_name") or "").strip(),
                start_date=_parse_date(r.get("start_date")) or "",
                end_date=_parse_date(r.get("end_date")),
                reason=str(r.get("reason") or "").strip() or None,
                is_medical=_parse_bool(r.get("is_medical")),
                granted_by_personal_number=str(r.get("granted_by_personal_number") or "").strip() or None,
                revoked=bool(_parse_bool(r.get("revoked"))),
                revoke_reason=str(r.get("revoke_reason") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "soldier_exemptions")
        ]

        system_settings = [
            ImportSystemSettingRow(
                source_row=r["_row"],
                key=str(r.get("key") or "").strip(),
                value_json=str(r["value_json"]).strip() if r.get("value_json") is not None else "",
            )
            for r in _sheet_rows(wb, "system_settings")
        ]
        rank_advancement_intervals = [
            ImportRankAdvancementIntervalRow(
                source_row=r["_row"],
                track=str(r.get("track") or "").strip(),
                rank=str(r.get("rank") or "").strip(),
                months_to_next=int(r.get("months_to_next") or 0),
                advance_on_career_entry=_parse_bool(r.get("advance_on_career_entry")),
            )
            for r in _sheet_rows(wb, "rank_advancement_intervals")
            if str(r.get("track") or "").strip()
        ]


        bug_reports = [
            ImportBugReportRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                reporter_personal_number=str(r.get("reporter_personal_number") or "").strip(),
                description=str(r.get("description") or "").strip(),
                severity=str(r.get("severity") or "").strip(),
                route=str(r.get("route") or "").strip(),
                status=str(r.get("status") or "").strip(),
                created_at=str(r.get("created_at") or "").strip() or None,
                nav_history_json=str(r.get("nav_history_json") or "").strip() or None,
                audit_snapshot_json=str(r.get("audit_snapshot_json") or "").strip() or None,
                user_snapshot_json=str(r.get("user_snapshot_json") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "bug_reports")
        ]

        range_locations = [
            ImportRangeLocationRow(
                source_row=r["_row"],
                name=str(r.get("name") or "").strip(),
                active=_parse_bool(r.get("active")),
            )
            for r in _sheet_rows(wb, "range_locations")
        ]

        range_events = [
            ImportRangeEventRow(
                source_row=r["_row"],
                hierarchy_node_name=str(r.get("hierarchy_node_name") or "").strip() or None,
                range_type=str(r.get("range_type") or "").strip(),
                date=_parse_date(r.get("date")) or "",
                range_location_name=str(r.get("range_location_name") or "").strip(),
                required_count=int(r.get("required_count") or 1),
                reserve_count=int(r.get("reserve_count") or 0),
                start_time=str(r.get("start_time") or "").strip() or None,
                end_time=str(r.get("end_time") or "").strip() or None,
                arrival_instructions=str(r.get("arrival_instructions") or "").strip() or None,
                contact_name=str(r.get("contact_name") or "").strip() or None,
                contact_phone=str(r.get("contact_phone") or "").strip() or None,
                notes=str(r.get("notes") or "").strip() or None,
                status=str(r.get("status") or "").strip() or None,
                responsible_duty_manager_personal_number=(
                    str(r.get("responsible_duty_manager_personal_number") or "").strip() or None
                ),
            )
            for r in _sheet_rows(wb, "range_events")
        ]

        range_assignments = [
            ImportRangeAssignmentRow(
                source_row=r["_row"],
                personal_number=str(r.get("personal_number") or "").strip(),
                full_name=str(r.get("full_name") or "").strip(),
                hierarchy_node_name=str(r.get("hierarchy_node_name") or "").strip() or None,
                range_type=str(r.get("range_type") or "").strip(),
                date=_parse_date(r.get("date")) or "",
                range_location_name=str(r.get("range_location_name") or "").strip(),
                is_reserve=_parse_bool(r.get("is_reserve")) or False,
                is_draft=_parse_bool(r.get("is_draft")) or False,
                attendance_status=str(r.get("attendance_status") or "").strip() or None,
                note=str(r.get("note") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "range_assignments")
        ]

        soldier_range_qualifications = [
            ImportSoldierRangeQualificationRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                range_type=str(r.get("range_type") or "").strip(),
                valid_until=_parse_date(r.get("valid_until")) or "",
            )
            for r in _sheet_rows(wb, "soldier_range_qualifications")
        ]

        range_excusal_requests = [
            ImportRangeExcusalRequestRow(
                source_row=r["_row"],
                id=str(r.get("id") or "").strip() or None,
                soldier_personal_number=str(r.get("soldier_personal_number") or "").strip(),
                requested_by_personal_number=str(r.get("requested_by_personal_number") or "").strip() or None,
                hierarchy_node_name=str(r.get("hierarchy_node_name") or "").strip() or None,
                range_type=str(r.get("range_type") or "").strip(),
                date=_parse_date(r.get("date")) or "",
                range_location_name=str(r.get("range_location_name") or "").strip(),
                reason=str(r.get("reason") or "").strip() or None,
                status=str(r.get("status") or "").strip(),
                decided_by_personal_number=str(r.get("decided_by_personal_number") or "").strip() or None,
                decision_note=str(r.get("decision_note") or "").strip() or None,
            )
            for r in _sheet_rows(wb, "range_excusal_requests")
        ]

        return ParsedImportData(
            soldiers=soldiers,
            duty_shifts=duty_shifts,
            assignments=assignments,
            duty_locations=duty_locations,
            hierarchy=hierarchy,
            duty_types=duty_types,
            shift_templates=shift_templates,
            exemption_types=exemption_types,
            swap_requests=swap_requests,
            exemption_requests=exemption_requests,
            soldier_field_updates=soldier_field_updates,
            soldier_enrollment_requests=soldier_enrollment_requests,
            personal_constraints=personal_constraints,
            soldier_exemptions=soldier_exemptions,
            system_settings=system_settings,
            bug_reports=bug_reports,
            range_locations=range_locations,
            range_events=range_events,
            range_assignments=range_assignments,
            soldier_range_qualifications=soldier_range_qualifications,
            range_excusal_requests=range_excusal_requests,
            parser_id=self.id,
            parser_warnings=warnings,
        )


register(V1StandardParser())
