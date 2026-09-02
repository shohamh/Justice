"""Bilingual (Hebrew/English) Excel export & import support.

Exports write Hebrew sheet names and Hebrew column headers, styled as real
Excel tables (header row + auto-filter + banded rows). Imports accept both
the Hebrew layout and the original English one: sheet names are resolved
through `HE_SHEETS` and header cells through `HE_HEADERS` (English headers
pass through unchanged).
"""
from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# ── Sheet names (canonical English → Hebrew) ────────────────────────────────

HE_SHEETS: dict[str, str] = {
    "soldiers": "חיילים",
    "duty_shifts": "משמרות",
    "assignments": "שיבוצים",
    "duty_locations": "מיקומי תורנויות",
    "hierarchy": "היררכיה",
    "duty_types": "סוגי תפקידים",
    "exemption_types": "סוגי פטורים",
    "shift_templates": "תבניות משמרות",
    "swap_requests": "בקשות חילוף",
    "exemption_requests": "בקשות פטור",
    "soldier_field_updates": "בקשות עדכון שדה",
    "soldier_enrollment_requests": "בקשות גיוס",
    "personal_constraints": "אילוצים אישיים",
    "soldier_exemptions": "פטורים",
    "system_settings": "הגדרות מערכת",
    "bug_reports": "דיווחי באגים",
    "range_locations": "מיקומי מטווח",
    "range_events": "ימי מטווח",
    "range_assignments": "שיבוצי מטווח",
    "soldier_range_qualifications": "כשירויות מטווח",
    "range_excusal_requests": "בקשות היעדרות",
    "rank_advancement_intervals": "מועדי קידום",
}

# ── Column headers (canonical English key → Hebrew) ─────────────────────────

HE_HEADERS: dict[str, str] = {
    "active": "פעיל",
    "applies_to_duty_types": "סוגי תפקידים",
    "approval_log": "יומן אישורים",
    "arrival_instructions": "הוראות הגעה",
    "attendance_status": "נוכחות",
    "audit_snapshot_json": "תמונת מצב ביקורת",
    "auto_roll": "חידוש אוטומטי",
    "auto_roll_until": "חידוש אוטומטי עד",
    "bahad1_graduate": "בהד 1",
    "base": "בסיס",
    "commander_approved_by_personal_number": "מס אישי מאשר מפקד",
    "commander_name": "שם מפקד",
    "commander_personal_number": "מס אישי מפקד",
    "contact_name": "איש קשר",
    "contact_phone": "טלפון איש קשר",
    "current_rank_since": "דרגה נוכחית מתאריך",
    "next_rank_date_overridden": "תאריך דרגה הבאה ידני",
    "profile_picture_url": "תמונת פרופיל",
    "food_type": "סוג מזון",
    "food_constraints": "מגבלות מזון",
    "telegram_chat_id": "מזהה צ'אט טלגרם",
    "telegram_username": "שם משתמש טלגרם",
    "telegram_is_verified": "טלגרם מאומת",
    "telegram_notifications_enabled": "התראות טלגרם פעילות",
    "telegram_verified_at": "טלגרם אומת בתאריך",
    "granted_at": "נוצר בתאריך",
    "requested_at": "נדרש בתאריך",
    "requesting_name": "שם פונה",
    "revoked_at": "בוטל בתאריך",
    "revoked_by_personal_number": "מס אישי מבטל",
    "soldier_name": "שם חייל",
    "unit_join_date": "תאריך הצטרפות ליחידה",
    "updated_at": "עודכן בתאריך",
    "rank_track": "מסלול דרגה",
    "assignment_reason_code": "קוד סיבת שיבוץ",
    "assignment_reason_text": "סיבת שיבוץ",
    "requires_weapon": "דורש נשק",
    "required_range_type": "סוג מטווח נדרש",
    "forbids_weapons": "אוסר נשק",
    "reserve_count_override": "עקיפת רזרבה",
    "months_to_next": "חודשים לדרגה הבאה",
    "advance_on_career_entry": "קידום בכניסה לקבע",
    "track": "מסלול",
    "covering_personal_number": "מס אישי מכסה",
    "covering_side_approved": "אישור צד מכסה",
    "created_at": "נוצר בתאריך",
    "date": "תאריך",
    "decided_by_personal_number": "מס אישי מחליט",
    "decision_note": "הערת החלטה",
    "description": "תיאור",
    "discharge_date": "תאריך שחרור",
    "duration_days": "משך בימים",
    "duty_date": "תאריך תפקיד",
    "duty_location_name": "מיקום",
    "duty_managers": "מנהלי תורנות",
    "duty_type_name": "סוג תפקיד",
    "eligible_units": "יחידות מותרות",
    "email": "אימייל",
    "end_date": "תאריך סיום",
    "end_time": "שעת סיום",
    "enlistment_date": "תאריך קליטה",
    "enrolled_at": "תאריך גיוס",
    "exemption_type_name": "סוג פטור",
    "field_name": "שם שדה",
    "files": "קבצים",
    "full_name": "שם מלא",
    "gender": "מין",
    "granted_by_personal_number": "מס אישי מעניק",
    "has_military_driving_license": "רישיון נהיגה צבאי",
    "hierarchy_node_name": "יחידה",
    "id": "מזהה",
    "instructions": "הנחיות",
    "is_career": "קבע",
    "is_commander_exemption": "פטור מפקד",
    "is_draft": "טיוטה",
    "is_external": "חיצוני",
    "is_global": "כללי",
    "is_medical": "רפואי",
    "is_officer": "קצין",
    "is_reserve": "מילואים",
    "key": "מפתח",
    "last_alal_date": "אלל אחרון",
    "last_mitvahim_date": "מתווה אחרון",
    "left_at": "תאריך עזיבה",
    "level": "רמה",
    "mandatory_end_date": "תאריך סיום חובה",
    "military_driving_license_expiry": "תוקף רישיון צבאי",
    "name": "שם",
    "nav_history_json": "היסטוריית ניווט",
    "new_value": "ערך חדש",
    "next_rank_date": "תאריך דרגה הבאה",
    "node_quotas": "מכסות יחידה",
    "note": "הערה",
    "notes": "הערות",
    "parent_name": "הורה",
    "personal_number": "מספר אישי",
    "phone": "טלפון",
    "previous_value": "ערך קודם",
    "range_location_name": "מטווח",
    "range_type": "סוג מטווח",
    "rank": "דרגה",
    "reason": "סיבה",
    "recurrence_type": "מחזוריות",
    "rejected_by_personal_number": "מס אישי דוחה",
    "reporter_personal_number": "מס אישי מדווח",
    "requested_by_personal_number": "מס אישי מבקש",
    "requested_node_name": "יחידה מבוקשת",
    "requester_side_approved": "אישור צד מבקש",
    "requesting_personal_number": "מס אישי פונה",
    "required_count": "נדרשים",
    "requirements_json": "דרישות",
    "reserve_count": "רזרבה",
    "reserve_minimum": "מינימום מילואים",
    "reserve_ratio": "יחס מילואים",
    "revoke_reason": "סיבת ביטול",
    "revoked": "בוטל",
    "route": "נתיב",
    "score_per_day": "ניקוד ליום",
    "severity": "חומרה",
    "soldier_personal_number": "מס אישי חייל",
    "start_date": "תאריך התחלה",
    "start_time": "שעת התחלה",
    "status": "סטטוס",
    "target_personal_number": "מס אישי יעד",
    "user_snapshot_json": "תמונת מצב משתמש",
    "valid_until": "תקף עד",
    "value_json": "ערך",
    "weekdays": "ימים בשבוע",
}

# Reverse map: Hebrew header → canonical English key.
_EN_HEADERS: dict[str, str] = {he: en for en, he in HE_HEADERS.items()}

# Reverse map: Hebrew sheet name → canonical English sheet name.
_EN_SHEETS: dict[str, str] = {he: en for en, he in HE_SHEETS.items()}


def canonical_sheet_name(name: str) -> str:
    """Canonical English sheet name for a workbook sheet (Hebrew or English)."""
    return _EN_SHEETS.get(name, name)


def hebrew_sheet_name(sheet_en: str) -> str:
    """Hebrew display name for a canonical English sheet (falls back to it)."""
    return HE_SHEETS.get(sheet_en, sheet_en)


def resolve_sheet_name(sheetnames: list[str], sheet_en: str) -> str | None:
    """Find the workbook sheet matching `sheet_en` by English or Hebrew name."""
    for candidate in (sheet_en, HE_SHEETS.get(sheet_en)):
        if candidate and candidate in sheetnames:
            return candidate
    return None


def hebrew_headers(sheet_en: str, headers_en: list[str]) -> list[str]:
    """Translate a canonical English header row to Hebrew (unknown → as-is)."""
    del sheet_en  # reserved for per-sheet overrides
    return [HE_HEADERS.get(h, h) for h in headers_en]


def canonical_headers(sheet_en: str, raw_headers: list[str]) -> list[str]:
    """Normalize a raw header row to canonical English keys.

    Hebrew headers map back through `HE_HEADERS`; English headers pass
    through lowercased (the historical convention).
    """
    del sheet_en  # reserved for per-sheet overrides
    out: list[str] = []
    for raw in raw_headers:
        header = (raw or "").strip().lower()
        out.append(_EN_HEADERS.get(header, header))
    return out


def apply_excel_table(ws: Worksheet, sheet_en: str) -> None:
    """Style the sheet's used range as a real Excel table: header row,
    filter buttons and banded (alternating-color) rows."""
    ws.sheet_view.rightToLeft = True
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return
    last_col_letter = ws.cell(row=1, column=max_col).column_letter
    table = Table(
        displayName=f"tbl_{sheet_en}",
        ref=f"A1:{last_col_letter}{max_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_bilingual_sheet(
    wb: openpyxl.Workbook,
    sheet_en: str,
    headers_en: list[str],
    rows: list[list[Any]],
) -> Worksheet:
    """Create a sheet with a Hebrew name, Hebrew headers, the given data rows,
    and Excel-table styling. Returns the worksheet."""
    ws = wb.create_sheet(hebrew_sheet_name(sheet_en))
    ws.append(hebrew_headers(sheet_en, headers_en))
    for row in rows:
        ws.append(row)
    apply_excel_table(ws, sheet_en)
    return ws


def finalize_bilingual_workbook(wb: openpyxl.Workbook) -> None:
    """Post-process an exported workbook in place:

    - rename each sheet to its Hebrew display name,
    - translate the header row (row 1) to Hebrew,
    - style the used range as a real Excel table (filter buttons + banded
      alternating-row colors) and flip the sheet to right-to-left.

    Writers keep appending canonical English names/headers; this single pass
    Hebraizes the whole workbook.
    """
    for ws in wb.worksheets:
        for sheet_en, sheet_he in HE_SHEETS.items():
            if ws.title != sheet_en:
                continue
            for cell in ws[1]:
                if cell.value is None:
                    continue
                key = str(cell.value).strip().lower()
                cell.value = HE_HEADERS.get(key, cell.value)
            ws.title = sheet_he
            apply_excel_table(ws, sheet_en)
            break
