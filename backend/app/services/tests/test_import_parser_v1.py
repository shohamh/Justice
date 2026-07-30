import openpyxl

from app.services.import_parsers.v1_standard import V1StandardParser


def _wb_with_duty_shifts_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_shifts")
    ws.append([
        "duty_type_name", "duty_location_name", "start_date", "end_date",
        "start_time", "end_time", "required_count", "node_quotas", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_duty_shifts_with_node_quotas():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5,
         "ענף פוקוס:2;ענף אלומות:3", ""],
    ])
    parser = V1StandardParser()
    data = parser.parse(wb)
    assert len(data.duty_shifts) == 1
    row = data.duty_shifts[0]
    assert row.duty_type_name == "שמירה"
    assert row.required_count == 5
    assert {(q.node_name, q.count) for q in row.node_quotas} == {
        ("ענף פוקוס", 2), ("ענף אלומות", 3),
    }


def test_parses_duty_shifts_without_node_quotas():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 3, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_shifts[0].node_quotas == []


def test_node_quotas_missing_colon_is_skipped_with_warning():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 3, "badformat", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_shifts[0].node_quotas == []
    assert any("badformat" in w for w in data.parser_warnings)


def test_node_quotas_non_integer_count_is_skipped_with_warning():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 3, "node:notanumber", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_shifts[0].node_quotas == []
    assert any("node:notanumber" in w for w in data.parser_warnings)


def test_node_quotas_mix_of_valid_and_invalid_entries():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 3,
         "ענף פוקוס:2;badformat;node:notanumber", ""],
    ])
    data = V1StandardParser().parse(wb)
    row = data.duty_shifts[0]
    assert {(q.node_name, q.count) for q in row.node_quotas} == {("ענף פוקוס", 2)}
    assert any("badformat" in w for w in data.parser_warnings)
    assert any("node:notanumber" in w for w in data.parser_warnings)


def test_detect_scores_high_for_known_sheet_names():
    wb = _wb_with_duty_shifts_sheet([])
    score = V1StandardParser().detect(wb)
    assert score >= 0.5


def test_detect_scores_low_for_unrelated_workbook():
    wb = openpyxl.Workbook()
    wb.active.title = "random_sheet"
    score = V1StandardParser().detect(wb)
    assert score < 0.5


def _wb_with_soldiers_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("soldiers")
    ws.append([
        "personal_number", "full_name", "rank", "gender", "is_officer",
        "hierarchy_node_name", "enrolled_at", "enlistment_date", "phone", "email",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_soldiers_sheet_row():
    wb = _wb_with_soldiers_sheet([
        ["12345", "ישראל ישראלי", "רב", "m", "false", "מדור א",
         "01.01.2022", "01.03.2020", "0500000000", "a@b.com"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.soldiers) == 1
    row = data.soldiers[0]
    assert row.personal_number == "12345"
    assert row.full_name == "ישראל ישראלי"
    assert row.is_officer is False
    assert row.hierarchy_node_name == "מדור א"
    assert row.enrolled_at == "2022-01-01"
    assert row.enlistment_date == "2020-03-01"


def test_soldiers_sheet_absent_gives_empty_list():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("duty_shifts").append([
        "duty_type_name", "duty_location_name", "start_date", "end_date",
        "start_time", "end_time", "required_count", "node_quotas", "notes",
    ])
    data = V1StandardParser().parse(wb)
    assert data.soldiers == []


def _wb_with_assignments_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("assignments")
    ws.append([
        "personal_number", "full_name", "duty_type_name", "duty_location_name",
        "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_assignments_sheet_row():
    wb = _wb_with_assignments_sheet([
        ["12345", "ישראל ישראלי", "שמירה", "שער ראשי",
         "15.06.2024", "16.06.2024", "20:00", "06:00", "true", "הערה"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.assignments) == 1
    row = data.assignments[0]
    assert row.personal_number == "12345"
    assert row.full_name == "ישראל ישראלי"
    assert row.duty_type_name == "שמירה"
    assert row.duty_location_name == "שער ראשי"
    assert row.start_date == "2024-06-15"
    assert row.end_date == "2024-06-16"
    assert row.start_time == "20:00"
    assert row.end_time == "06:00"
    assert row.is_reserve is True
    assert row.notes == "הערה"


def test_assignments_sheet_does_not_produce_synthetic_duty_shifts():
    wb = _wb_with_assignments_sheet([
        ["12345", "ישראל ישראלי", "שמירה", "שער ראשי",
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_shifts == []
    assert not any("assignments" in w for w in data.parser_warnings)


def test_assignments_sheet_absent_gives_empty_list():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    data = V1StandardParser().parse(wb)
    assert data.assignments == []


def test_assignments_and_duty_shifts_both_present_both_parsed():
    wb = _wb_with_assignments_sheet([
        ["12345", "ישראל ישראלי", "שמירה", "שער ראשי",
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    ws = wb.create_sheet("duty_shifts")
    ws.append([
        "duty_type_name", "duty_location_name", "start_date", "end_date",
        "start_time", "end_time", "required_count", "node_quotas", "notes",
    ])
    ws.append(["שמירה", "שער ראשי", "15.06.2024", "16.06.2024", "", "", 2, "", ""])

    data = V1StandardParser().parse(wb)
    assert len(data.assignments) == 1
    assert len(data.duty_shifts) == 1


def _wb_with_hierarchy_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("hierarchy")
    ws.append([
        "name", "level", "parent_name", "commander_personal_number", "commander_name", "duty_managers",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_hierarchy_sheet_with_duty_managers():
    wb = _wb_with_hierarchy_sheet([
        ["מדור א", "group", "יחידה ראשית", "12345", "ישראל ישראלי", "12345:ישראל ישראלי;23456:משה כהן"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.hierarchy) == 1
    row = data.hierarchy[0]
    assert row.name == "מדור א"
    assert row.level == "group"
    assert row.parent_name == "יחידה ראשית"
    assert row.commander_personal_number == "12345"
    assert row.commander_name == "ישראל ישראלי"
    assert row.duty_manager_refs == ["12345:ישראל ישראלי", "23456:משה כהן"]


def test_hierarchy_malformed_duty_manager_entry_produces_warning_not_error():
    wb = _wb_with_hierarchy_sheet([
        ["מדור ב", "group", "", "", "", "not-a-valid-entry"],
    ])
    data = V1StandardParser().parse(wb)
    assert data.hierarchy[0].duty_manager_refs == []
    assert any("מדור ב" in w or "שורה 2" in w for w in data.parser_warnings)


def _wb_with_duty_locations_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_locations")
    ws.append(["name", "base", "active"])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_duty_locations_sheet():
    wb = _wb_with_duty_locations_sheet([
        ["בסיס א", "base_a", True],
        ["בסיס ב", "base_b", False],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_locations) == 2
    assert data.duty_locations[0].name == "בסיס א"
    assert data.duty_locations[0].base == "base_a"
    assert data.duty_locations[0].active is True
    assert data.duty_locations[1].name == "בסיס ב"
    assert data.duty_locations[1].active is False


def test_duty_locations_sheet_absent_gives_empty_list():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_locations == []


def test_parses_duty_locations_with_optional_fields():
    wb = _wb_with_duty_locations_sheet([
        ["בסיס א", None, None],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_locations) == 1
    row = data.duty_locations[0]
    assert row.name == "בסיס א"
    assert row.base is None
    assert row.active is None


def _wb_with_exemption_types_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("exemption_types")
    ws.append(["name", "description", "is_global", "is_medical", "is_commander_exemption", "applies_to_duty_types"])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_exemption_types_sheet():
    wb = _wb_with_exemption_types_sheet([
        ["פטור בריאות", "health reason", True, True, False, ""],
        ["פטור שמירות", "security reason", False, False, False, "שמירה,טיול"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.exemption_types) == 2
    assert data.exemption_types[0].name == "פטור בריאות"
    assert data.exemption_types[0].is_global is True
    assert data.exemption_types[0].is_medical is True
    assert data.exemption_types[0].applies_to_duty_type_names == []
    assert data.exemption_types[1].name == "פטור שמירות"
    assert data.exemption_types[1].applies_to_duty_type_names == ["שמירה", "טיול"]


def test_exemption_types_sheet_absent_gives_empty_list():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.exemption_types == []


def test_parses_exemption_types_with_whitespace_in_applies_to_list():
    wb = _wb_with_exemption_types_sheet([
        ["פטור שמירות", "", False, False, False, "שמירה , טיול , ביקור"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.exemption_types) == 1
    row = data.exemption_types[0]
    assert row.applies_to_duty_type_names == ["שמירה", "טיול", "ביקור"]


def _wb_with_duty_types_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_types")
    ws.append([
        # Column is `eligible_units` in the real spreadsheet layout — the
        # parser reads r.get("eligible_units") into the ImportDutyTypeRow
        # field named `eligible_unit_names`. Keep the header name matching
        # the actual sheet, not the schema attribute name.
        "name", "score_per_day", "description", "active", "reserve_ratio",
        "reserve_minimum", "is_external", "contact_name", "contact_phone",
        "start_time", "end_time", "instructions", "eligible_units", "requirements_json",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_duty_types_sheet():
    wb = _wb_with_duty_types_sheet([
        [
            "שמירה", "1.50", "תיאור", "true", "0.200",
            "2", "false", "דני", "050-1234567",
            "20:00", "06:00", "הצטיידות", "מדור א, מדור ב", '{"min_rank": 1}',
        ],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_types) == 1
    row = data.duty_types[0]
    assert row.name == "שמירה"
    assert row.score_per_day == "1.50"
    assert row.description == "תיאור"
    assert row.active is True
    assert row.reserve_ratio == "0.200"
    assert row.reserve_minimum == 2
    assert row.is_external is False
    assert row.contact_name == "דני"
    assert row.contact_phone == "050-1234567"
    assert row.start_time == "20:00"
    assert row.end_time == "06:00"
    assert row.instructions == "הצטיידות"
    # Regression test: the spreadsheet column is `eligible_units`, not
    # `eligible_unit_names` — the parser must read from the real column name.
    assert row.eligible_unit_names == ["מדור א", "מדור ב"]
    assert row.requirements_json == '{"min_rank": 1}'



def test_duty_types_numeric_fields_stay_as_raw_strings_or_ints():
    wb = _wb_with_duty_types_sheet([
        [
            "שמירה", "1.50", "", "", "0.200",
            "2", "", "", "",
            "", "", "", "", "",
        ],
    ])
    data = V1StandardParser().parse(wb)
    row = data.duty_types[0]
    # score_per_day and reserve_ratio are kept as raw strings (not Decimal)
    # at this parsing stage — conversion happens later in validation.
    assert row.score_per_day == "1.50"
    assert isinstance(row.score_per_day, str)
    assert row.reserve_ratio == "0.200"
    assert isinstance(row.reserve_ratio, str)
    # reserve_minimum is parsed as an int at this stage.
    assert row.reserve_minimum == 2
    assert isinstance(row.reserve_minimum, int)


def test_duty_types_sheet_absent_gives_empty_list():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.duty_types == []


def test_parses_duty_types_with_optional_fields_none():
    wb = _wb_with_duty_types_sheet([
        ["טיול", "2.00", None, None, None, None, None, None, None, None, None, None, None, None],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_types) == 1
    row = data.duty_types[0]
    assert row.name == "טיול"
    assert row.score_per_day == "2.00"
    assert row.description is None
    assert row.active is None
    assert row.reserve_ratio is None
    assert row.reserve_minimum is None
    assert row.is_external is None
    assert row.contact_name is None
    assert row.contact_phone is None
    assert row.start_time is None
    assert row.end_time is None
    assert row.instructions is None
    assert row.eligible_unit_names == []
    assert row.requirements_json is None


def test_parses_duty_types_with_whitespace_in_eligible_units():
    wb = _wb_with_duty_types_sheet([
        ["שמירה", "1.50", "", False, "", 0, False, "", "", "", "", "", "יחידה א , יחידה ב , יחידה ג", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_types) == 1
    row = data.duty_types[0]
    assert row.eligible_unit_names == ["יחידה א", "יחידה ב", "יחידה ג"]


def test_parses_multiple_duty_types():
    wb = _wb_with_duty_types_sheet([
        ["שמירה", "1.50", "guard", True, "0.30", 5, False, "עמית", "0500000000", "", "", "", "יחידה א", ""],
        ["טיול", "2.00", "tour", False, None, None, True, "", "", "10:00", "18:00", "fun", "", '{"skill":"navigation"}'],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.duty_types) == 2
    assert data.duty_types[0].name == "שמירה"
    assert data.duty_types[1].name == "טיול"


def test_duty_types_reserve_minimum_zero_is_not_lost():
    # A literal 0 in the reserve_minimum cell is falsy in Python, so a naive
    # `if r.get("reserve_minimum") else None` would incorrectly turn a real
    # zero into None. Must distinguish "cell present and zero" from "cell
    # blank" using string-emptiness, not truthiness.
    wb = _wb_with_duty_types_sheet([
        [
            "שמירה", "1.50", "", "", "0.200",
            0, "", "", "",
            "", "", "", "", "",
        ],
    ])
    data = V1StandardParser().parse(wb)
    row = data.duty_types[0]
    assert row.reserve_minimum == 0


def _wb_with_shift_templates(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("shift_templates")
    ws.append([
        "name", "duty_type_name", "duty_location_name", "recurrence_type", "weekdays",
        "start_time", "end_time", "required_count", "auto_roll", "auto_roll_until",
        "duration_days", "notes", "eligible_units",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_shift_templates_sheet_row():
    wb = _wb_with_shift_templates([
        ["שמירה לילה", "שמירה", "שער ראשי", "weekly", "1,3",
         "20:00", "06:00", 2, "true", "31.12.2026", 1, "הערה", "מדור א"],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.shift_templates) == 1
    row = data.shift_templates[0]
    assert row.name == "שמירה לילה"
    assert row.duty_type_name == "שמירה"
    assert row.duty_location_name == "שער ראשי"
    assert row.recurrence_type == "weekly"
    assert row.weekdays == [1, 3]
    assert row.start_time == "20:00"
    assert row.end_time == "06:00"
    assert row.required_count == 2
    assert row.auto_roll is True
    assert row.auto_roll_until == "2026-12-31"
    assert row.duration_days == 1
    assert row.notes == "הערה"
    assert row.eligible_unit_names == ["מדור א"]


def test_shift_templates_sheet_absent_gives_empty_list():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    data = V1StandardParser().parse(wb)
    assert data.shift_templates == []


def test_shift_templates_row_defaults():
    # A blank cell now parses to None at the parser level for these four fields
    # (recurrence_type/required_count/auto_roll/duration_days), rather than being
    # coerced to a hard default here. That lets a blank cell on an UPDATE row
    # mean "leave the existing value unchanged"; the create-time default
    # (weekdays/1/False/1) is applied only for brand-new rows, in
    # confirm_session's shift_templates loop.
    wb = _wb_with_shift_templates([
        ["שמירה", "שמירה", "שער ראשי", "", "", "", "", "", "", "", "", "", ""],
    ])
    data = V1StandardParser().parse(wb)
    row = data.shift_templates[0]
    assert row.recurrence_type is None
    assert row.weekdays == []
    assert row.required_count is None
    assert row.auto_roll is None
    assert row.duration_days is None


def _wb_with_system_settings_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("system_settings")
    ws.append(["key", "value_json"])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_system_settings_sheet():
    wb = _wb_with_system_settings_sheet([
        ["algorithm.max_duties_per_window", "8"],
        ["telegram.enabled", "true"],
        ["registration.default_role", '"soldier"'],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.system_settings) == 3
    assert data.system_settings[0].key == "algorithm.max_duties_per_window"
    assert data.system_settings[0].value_json == "8"
    assert data.system_settings[1].value_json == "true"
    assert data.system_settings[2].value_json == '"soldier"'


def test_system_settings_sheet_absent_gives_empty_list():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.system_settings == []


def _wb_with_bug_reports_sheet(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("bug_reports")
    ws.append([
        "id", "reporter_personal_number", "description", "severity", "route", "status",
        "created_at", "nav_history_json", "audit_snapshot_json", "user_snapshot_json",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_parses_bug_reports_sheet():
    wb = _wb_with_bug_reports_sheet([
        ["", "1234567", "הכפתור לא עובד", "medium", "/planning/export", "open",
         "", "", "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert len(data.bug_reports) == 1
    row = data.bug_reports[0]
    assert row.id is None
    assert row.reporter_personal_number == "1234567"
    assert row.description == "הכפתור לא עובד"
    assert row.severity == "medium"
    assert row.route == "/planning/export"
    assert row.status == "open"
    assert row.nav_history_json is None


def test_parses_bug_reports_sheet_with_id_and_json_columns():
    wb = _wb_with_bug_reports_sheet([
        ["11111111-1111-1111-1111-111111111111", "1234567", "תקלה", "high",
         "/x", "resolved", "2026-01-01T00:00:00+00:00",
         '[{"path": "/a"}]', '[{"action": "x"}]', '{"role": "soldier"}'],
    ])
    data = V1StandardParser().parse(wb)
    row = data.bug_reports[0]
    assert row.id == "11111111-1111-1111-1111-111111111111"
    assert row.nav_history_json == '[{"path": "/a"}]'
    assert row.audit_snapshot_json == '[{"action": "x"}]'
    assert row.user_snapshot_json == '{"role": "soldier"}'


def test_bug_reports_sheet_absent_gives_empty_list():
    wb = _wb_with_duty_shifts_sheet([
        ["שמירה", "בסיס א", "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    data = V1StandardParser().parse(wb)
    assert data.bug_reports == []
