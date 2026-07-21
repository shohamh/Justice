from __future__ import annotations

import io
import uuid
from datetime import date as date_type
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import select

from app.db.models import DutyManagerScope, DutyLocation, DutyShiftNodeQuota, Soldier
from app.services.duty_config import create_duty_type
import app.services.import_parsers.v1_standard  # noqa: F401  (registers "v1_standard" parser)
from app.services.import_sessions import (
    ImportSessionError,
    cancel_session,
    confirm_session,
    create_session,
    mark_done,
    reparse_session,
    set_selections,
)
from tests.helpers import create_node, create_soldier


def _uid() -> str:
    return uuid.uuid4().hex[:8]


def _wb_with_duty_shifts(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_shifts")
    ws.append([
        "duty_type_name", "duty_location_name", "start_date", "end_date",
        "start_time", "end_time", "required_count", "node_quotas", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_soldiers(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("soldiers")
    ws.append([
        "personal_number", "full_name", "rank", "gender", "is_officer",
        "hierarchy_node_name", "enrolled_at", "enlistment_date", "phone", "email",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_create_session_admin_resolvable_duty_shift(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    assert sess.status == "draft"
    row = sess.parsed_state["duty_shifts"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_type_id"] == str(dt.id)
    assert row["resolved_duty_location_id"] == str(loc.id)


def test_create_session_dm_out_of_scope_quota_node(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()

    in_scope_node = create_node(admin_session, level="branch", name=f"in_{_uid()}")
    out_of_scope_node = create_node(admin_session, level="branch", name=f"out_{_uid()}")
    dm = create_soldier(admin_session, personal_number=f"dm_{_uid()}", role="duty_manager")
    admin_session.add(DutyManagerScope(duty_manager_id=dm.id, hierarchy_node_id=in_scope_node.id))
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5,
         f"{out_of_scope_node.name}:2", ""],
    ])

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=dm, parser_id="v1_standard",
    )

    row = sess.parsed_state["duty_shifts"][0]
    assert row["action"] == "out_of_scope"


def test_create_session_unresolved_duty_type_errors(admin_session):
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        ["no_such_duty_type", loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    row = sess.parsed_state["duty_shifts"][0]
    assert row["action"] == "error"
    assert any("no_such_duty_type" in e for e in row["errors"])


def test_create_session_soldier_unresolved_hierarchy_node_errors(admin_session):
    wb = _wb_with_soldiers([
        ["1234567", "Some Soldier", "", "", "", "no_such_node", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    row = sess.parsed_state["soldiers"][0]
    assert row["action"] == "error"
    assert any("no_such_node" in e for e in row["errors"])


def test_reparse_session_flips_error_to_new_after_duty_type_created(admin_session):
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    dt_name = f"dt_{_uid()}"
    wb = _wb_with_duty_shifts([
        [dt_name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()
    assert sess.parsed_state["duty_shifts"][0]["action"] == "error"

    create_duty_type(admin_session, name=dt_name, score_per_day=Decimal("1.00"))
    admin_session.commit()

    reparsed = reparse_session(admin_session, session_id=sess.id, actor=admin)
    assert reparsed.parsed_state["duty_shifts"][0]["action"] == "new"


def test_reparse_session_non_draft_raises(admin_session):
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    sess.status = "confirmed"
    admin_session.commit()

    with pytest.raises(ImportSessionError, match="only draft sessions"):
        reparse_session(admin_session, session_id=sess.id, actor=admin)


def test_reparse_session_not_found_raises(admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    with pytest.raises(ImportSessionError, match="not found"):
        reparse_session(admin_session, session_id=uuid.uuid4(), actor=admin)


def test_set_selections_persists(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    selections = {"duty_shifts": {"1": "skip"}}
    updated = set_selections(admin_session, session_id=sess.id, selections=selections)
    assert updated.user_selections == selections

    admin_session.commit()
    admin_session.refresh(sess)
    assert sess.user_selections == selections


def test_confirm_session_mixed_actions(admin_session):
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.commit()

    wb = _wb_with_soldiers([
        ["1111111", "New Soldier", "", "", "", "", "", "", "", ""],
        ["2222222", "Skip Soldier", "", "", "", "", "", "", "", ""],
        ["no_such_node_soldier", "Error Soldier", "", "", "", node.name + "_bogus", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    rows = sess.parsed_state["soldiers"]
    assert rows[0]["action"] == "new"
    assert rows[1]["action"] == "new"
    assert rows[2]["action"] == "error"

    # User selects "skip" for row 2 (the second row, effective row number = its source_row)
    row2_num = str(rows[1]["row"])
    set_selections(admin_session, session_id=sess.id, selections={"soldiers": {row2_num: "skip"}})
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 1
    assert result["skipped"] == 2
    assert result["updated"] == 0

    admin_session.refresh(sess)
    assert sess.status == "confirmed"
    assert sess.confirmed_at is not None
    assert len(sess.created_links["soldiers"]) == 1

    created_id = sess.created_links["soldiers"][0]
    created_soldier = admin_session.get(Soldier, uuid.UUID(created_id))
    assert created_soldier is not None
    assert created_soldier.personal_number == "1111111"


def test_confirm_session_duty_shift_with_node_quota_writes_quota(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, f"{node.name}:3", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 1
    assert result["skipped"] == 0

    shift_id = uuid.UUID(sess.created_links["duty_shifts"][0])
    quotas = admin_session.execute(
        select(DutyShiftNodeQuota).where(DutyShiftNodeQuota.duty_shift_id == shift_id)
    ).scalars().all()
    assert len(quotas) == 1
    assert quotas[0].hierarchy_node_id == node.id
    assert quotas[0].count == 3


def test_confirm_session_duty_shift_quota_failure_leaves_no_partial_shift(admin_session):
    # Regression test: shift creation succeeds and is flushed, but the
    # subsequent set_shift_quotas call fails (here, because the quota's
    # hierarchy node was deleted after parsing but before confirmation,
    # even though parsed_state still marks it "resolved": true). The row
    # must have zero effect overall — no orphaned DutyShift persisted
    # without its quotas.
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, f"{node.name}:3", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    assert sess.parsed_state["duty_shifts"][0]["node_quotas"][0]["resolved"] is True

    # Delete the node after parsing so parsed_state still says resolved=True
    # but set_shift_quotas will fail with "hierarchy node not found".
    admin_session.delete(admin_session.get(type(node), node.id))
    admin_session.commit()

    from app.db.models import DutyShift

    count_before = admin_session.execute(select(DutyShiftNodeQuota)).scalars().all()
    shifts_before = admin_session.execute(select(DutyShift)).scalars().all()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert result["errors"][0]["type"] == "duty_shifts"

    shifts_after = admin_session.execute(select(DutyShift)).scalars().all()
    assert len(shifts_after) == len(shifts_before)

    quotas_after = admin_session.execute(select(DutyShiftNodeQuota)).scalars().all()
    assert len(quotas_after) == len(count_before)

    assert sess.created_links["duty_shifts"] == []


def test_confirm_session_non_draft_raises(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    sess.status = "confirmed"
    admin_session.commit()

    with pytest.raises(ImportSessionError, match="only draft sessions"):
        confirm_session(admin_session, session_id=sess.id, actor=admin)


def test_cancel_session_draft(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    cancelled = cancel_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert cancelled.status == "cancelled"
    assert cancelled.cancelled_at is not None


def test_mark_done_requires_confirmed(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    with pytest.raises(ImportSessionError, match="confirmed"):
        mark_done(admin_session, session_id=sess.id, actor=admin)

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    assert result["created"] == 1

    done = mark_done(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    assert done.status == "done"


def test_reparse_resolves_duty_type_via_by_name_mapping(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    # Excel uses a different name than the DB
    wb = _wb_with_duty_shifts([
        ["excel_alias", loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    # Initially unresolved
    assert sess.parsed_state["duty_shifts"][0]["action"] == "error"

    # Apply by_name mapping
    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {
            "duty_type": {"by_name": {"excel_alias": str(dt.id)}}
        }
    })
    admin_session.commit()

    sess = reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_shifts"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_type_id"] == str(dt.id)


def test_reparse_by_row_overrides_by_name_for_duty_type(admin_session):
    dt_name = create_duty_type(admin_session, name=f"dt_name_{_uid()}", score_per_day=Decimal("1.00"))
    dt_row  = create_duty_type(admin_session, name=f"dt_row_{_uid()}",  score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        ["excel_alias", loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    row_number = sess.parsed_state["duty_shifts"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {
            "duty_type": {
                "by_name": {"excel_alias": str(dt_name.id)},
                "by_row":  {f"duty_shifts:{row_number}": str(dt_row.id)},
            }
        }
    })
    admin_session.commit()

    sess = reparse_session(admin_session, session_id=sess.id, actor=admin)
    assert sess.parsed_state["duty_shifts"][0]["resolved_duty_type_id"] == str(dt_row.id)


def test_reparse_resolves_hierarchy_node_via_by_name_mapping(admin_session):
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.commit()

    wb = _wb_with_soldiers([
        ["1234567", "Test Soldier", "", "", "", "excel_node_alias", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    assert sess.parsed_state["soldiers"][0]["action"] == "error"

    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {
            "hierarchy_node": {"by_name": {"excel_node_alias": str(node.id)}}
        }
    })
    admin_session.commit()

    sess = reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["soldiers"][0]
    assert row["action"] == "new"
    assert row["hierarchy_node_id"] == str(node.id)


def test_reparse_resolves_quota_node_via_by_row_mapping(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "excel_quota_node:3", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    initial_quotas = sess.parsed_state["duty_shifts"][0]["node_quotas"]
    assert initial_quotas[0]["resolved"] is False
    row_number = sess.parsed_state["duty_shifts"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {
            "hierarchy_node": {
                "by_row": {f"duty_shifts:{row_number}:excel_quota_node": str(node.id)}
            }
        }
    })
    admin_session.commit()

    sess = reparse_session(admin_session, session_id=sess.id, actor=admin)
    quotas = sess.parsed_state["duty_shifts"][0]["node_quotas"]
    assert quotas[0]["resolved"] is True
    assert quotas[0]["node_id"] == str(node.id)


def test_reparse_bad_mapped_uuid_falls_back_to_name_lookup(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    assert sess.parsed_state["duty_shifts"][0]["action"] == "new"  # baseline: resolves by name

    bad_uuid = str(uuid.uuid4())  # does not exist in DB
    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {
            "duty_type": {"by_name": {dt.name: bad_uuid}}
        }
    })
    admin_session.commit()

    sess = reparse_session(admin_session, session_id=sess.id, actor=admin)
    # Bad UUID → not found in DB → fallback to name lookup → still resolves correctly
    assert sess.parsed_state["duty_shifts"][0]["action"] == "new"
    assert sess.parsed_state["duty_shifts"][0]["resolved_duty_type_id"] == str(dt.id)


def test_soldier_fallback_matches_by_full_name_when_personal_number_unknown(admin_session):
    existing = create_soldier(admin_session, personal_number=f"old_pn_{_uid()}")
    admin_session.commit()

    wb = _wb_with_soldiers([
        [f"new_pn_{_uid()}", existing.full_name, "", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["soldiers"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(existing.id)
    assert any("שם" in w for w in row["warnings"])


def test_soldier_fallback_ambiguous_full_name_errors(admin_session):
    from app.auth.password import hash_password
    from app.db.models import Soldier

    dup_name = f"Dup Name {_uid()}"
    s1 = Soldier(personal_number=f"pn1_{_uid()}", full_name=dup_name, password_hash=hash_password("x"))
    s2 = Soldier(personal_number=f"pn2_{_uid()}", full_name=dup_name, password_hash=hash_password("x"))
    admin_session.add_all([s1, s2])
    admin_session.commit()

    wb = _wb_with_soldiers([
        [f"unknown_pn_{_uid()}", dup_name, "", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["soldiers"][0]
    assert row["action"] == "error"
    assert any("חד משמעי" in e for e in row["errors"])


def test_soldier_fallback_updates_personal_number_on_confirm(admin_session):
    existing = create_soldier(admin_session, personal_number=f"old_pn_{_uid()}")
    admin_session.commit()
    new_pn = f"new_pn_{_uid()}"

    wb = _wb_with_soldiers([
        [new_pn, existing.full_name, "", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    assert result["updated"] == 1

    admin_session.refresh(existing)
    assert existing.personal_number == new_pn


def _wb_with_assignments(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("assignments")
    ws.append([
        "personal_number", "full_name", "duty_type_name", "duty_location_name",
        "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_duty_shifts_and_assignments(duty_shift_rows, assignment_rows):
    wb = _wb_with_duty_shifts(duty_shift_rows)
    ws = wb.create_sheet("assignments")
    ws.append([
        "personal_number", "full_name", "duty_type_name", "duty_location_name",
        "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes",
    ])
    for r in assignment_rows:
        ws.append(r)
    return wb


def test_assignment_matches_existing_shift(admin_session):
    from app.db.models import DutyShift

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=2,
    )
    admin_session.add(shift)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "00:00", "23:59", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_shift_id"] == str(shift.id)
    assert row["matched_session_row"] is None


def test_assignment_matches_session_duty_shifts_row(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_duty_shifts_and_assignments(
        [[dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 2, "", ""]],
        [[soldier.personal_number, soldier.full_name, dt.name, loc.name,
          "15.06.2024", "16.06.2024", "", "", "false", ""]],
    )
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )

    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_shift_id"] is None
    assert row["matched_session_row"] == sess.parsed_state["duty_shifts"][0]["row"]


def test_assignment_full_name_mismatch_errors(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, "Wrong Name", dt.name, loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "error"
    assert any("שם מלא" in e for e in row["errors"])


def test_assignment_personal_number_unknown_falls_back_to_full_name(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_assignments([
        [f"unknown_{_uid()}", soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "error"  # no matching shift exists for this dt/loc/dates
    # Soldier itself resolved via fallback despite no shift match:
    assert row["resolved_soldier_id"] == str(soldier.id)
    assert any("נמצא לפי שם" in w for w in row["warnings"])


def test_assignment_ambiguous_full_name_errors(admin_session):
    from app.auth.password import hash_password
    from app.db.models import Soldier

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    dup_name = f"Dup {_uid()}"
    s1 = Soldier(personal_number=f"pn1_{_uid()}", full_name=dup_name, password_hash=hash_password("x"))
    s2 = Soldier(personal_number=f"pn2_{_uid()}", full_name=dup_name, password_hash=hash_password("x"))
    admin_session.add_all([s1, s2])
    admin_session.commit()

    wb = _wb_with_assignments([
        [f"unknown_{_uid()}", dup_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "error"
    assert any("חד משמעי" in e for e in row["errors"])


def test_assignment_duplicate_is_skipped(admin_session):
    from app.db.models import DutyAssignment, DutyShift

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=2,
    )
    admin_session.add(shift)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.add(DutyAssignment(
        soldier_id=soldier.id, duty_type_id=dt.id, duty_location_id=loc.id,
        duty_shift_id=shift.id, start_date=shift.start_date, end_date=shift.end_date,
    ))
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "00:00", "23:59", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "skip"


def test_assignment_over_capacity_warns_but_allows(admin_session):
    from app.db.models import DutyAssignment, DutyShift

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=1,
    )
    admin_session.add(shift)
    admin_session.flush()
    already_assigned = create_soldier(admin_session, personal_number=f"sol_a_{_uid()}")
    admin_session.add(DutyAssignment(
        soldier_id=already_assigned.id, duty_type_id=dt.id, duty_location_id=loc.id,
        duty_shift_id=shift.id, start_date=shift.start_date, end_date=shift.end_date,
    ))
    new_soldier = create_soldier(admin_session, personal_number=f"sol_b_{_uid()}")
    admin_session.commit()

    wb = _wb_with_assignments([
        [new_soldier.personal_number, new_soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "00:00", "23:59", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "new"
    assert any("1/1" in w for w in row["warnings"])


def test_confirm_session_creates_assignment_against_existing_shift(admin_session):
    from app.db.models import DutyAssignment, DutyShift

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=2,
    )
    admin_session.add(shift)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "00:00", "23:59", "true", "note"],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 1
    assert result["errors"] == []
    assert len(sess.created_links["assignments"]) == 1

    created = admin_session.get(DutyAssignment, uuid.UUID(sess.created_links["assignments"][0]))
    assert created is not None
    assert created.soldier_id == soldier.id
    assert created.duty_shift_id == shift.id
    assert created.is_reserve is True
    assert created.notes == "note"


def test_confirm_session_creates_assignment_against_session_created_shift(admin_session):
    from app.db.models import DutyAssignment

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_duty_shifts_and_assignments(
        [[dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 2, "", ""]],
        [[soldier.personal_number, soldier.full_name, dt.name, loc.name,
          "15.06.2024", "16.06.2024", "", "", "false", ""]],
    )
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 2  # 1 duty_shift + 1 assignment
    assert result["errors"] == []
    created_shift_id = uuid.UUID(sess.created_links["duty_shifts"][0])
    created_assignment = admin_session.get(
        DutyAssignment, uuid.UUID(sess.created_links["assignments"][0])
    )
    assert created_assignment.duty_shift_id == created_shift_id


def test_confirm_session_assignment_matched_shift_skipped_errors_gracefully(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()

    wb = _wb_with_duty_shifts_and_assignments(
        [[dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 2, "", ""]],
        [[soldier.personal_number, soldier.full_name, dt.name, loc.name,
          "15.06.2024", "16.06.2024", "", "", "false", ""]],
    )
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    shift_row_num = sess.parsed_state["duty_shifts"][0]["row"]
    set_selections(admin_session, session_id=sess.id, selections={
        "duty_shifts": {str(shift_row_num): "skip"},
    })
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 0
    assert len(result["errors"]) == 1
    assert result["errors"][0]["type"] == "assignments"
    assert sess.created_links["assignments"] == []


def _wb_with_duty_types(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_types")
    ws.append([
        "name", "score_per_day", "description", "active", "reserve_ratio", "reserve_minimum",
        "is_external", "contact_name", "contact_phone", "start_time", "end_time",
        "instructions", "eligible_units", "requirements_json",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_exemption_types(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("exemption_types")
    ws.append([
        "name", "description", "is_global", "is_medical", "is_commander_exemption",
        "applies_to_duty_types",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_duty_locations(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_locations")
    ws.append(["name", "base", "active"])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_hierarchy(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("hierarchy")
    ws.append([
        "name", "level", "parent_name", "commander_personal_number",
        "commander_name", "duty_manager_refs",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_duty_type_field_override_changes_resolved_score(admin_session):
    wb = _wb_with_duty_types([
        ["שמירה", "1.00", "", "true", "", "", "false", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_types"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_types": {str(row_num): {"score_per_day": "2.50"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_types"][0]
    assert row["score_per_day"] == "2.50"


def test_duty_type_field_override_invalid_value_produces_row_error(admin_session):
    wb = _wb_with_duty_types([
        ["שמירה", "1.00", "", "true", "", "", "false", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_types"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_types": {str(row_num): {"score_per_day": "not-a-number"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_types"][0]
    assert row["action"] == "error"
    assert any("ניקוד ליום לא תקין" in e for e in row["errors"])


def test_exemption_type_field_override_changes_resolved_flag(admin_session):
    wb = _wb_with_exemption_types([
        ["פטור", "", "false", "false", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["exemption_types"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"exemption_types": {str(row_num): {"is_medical": True}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["exemption_types"][0]
    assert row["is_medical"] is True


def test_duty_type_field_override_bypasses_eligible_unit_resolution(admin_session):
    wb = _wb_with_duty_types([
        ["שמירה", "1.00", "", "true", "", "", "false", "", "", "", "", "", "לא קיים", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_types"][0]["row"]
    # Without the override, "לא קיים" fails to resolve and the row errors.
    assert sess.parsed_state["duty_types"][0]["action"] == "error"

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_types": {str(row_num): {"resolved_eligible_node_ids": []}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_types"][0]
    assert row["action"] == "new"
    assert row["resolved_eligible_node_ids"] == []


def test_exemption_type_field_override_bypasses_applies_to_resolution(admin_session):
    wb = _wb_with_exemption_types([
        ["פטור", "", "false", "false", "false", "לא קיים"],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["exemption_types"][0]["row"]
    assert sess.parsed_state["exemption_types"][0]["action"] == "error"

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"exemption_types": {str(row_num): {"resolved_duty_type_ids": []}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["exemption_types"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_type_ids"] == []


def test_duty_location_field_override_changes_base(admin_session):
    wb = _wb_with_duty_locations([
        [f"loc_{_uid()}", "Original Base", "true"],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_locations"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_locations": {str(row_num): {"base": "Overridden Base"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_locations"][0]
    assert row["base"] == "Overridden Base"


def test_confirm_session_new_duty_type_preserves_requirements(admin_session):
    from app.db.models import DutyType as DutyTypeModel

    wb = _wb_with_duty_types([
        ["שמירה", "1.00", "", "true", "", "", "false", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_types"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_types": {str(row_num): {"requirements": {"officers_allowed": False}}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 1
    created = admin_session.execute(
        select(DutyTypeModel).where(DutyTypeModel.name == "שמירה")
    ).scalar_one()
    assert created.requirements == {"officers_allowed": False}


def _wb_with_shift_templates(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("shift_templates")
    ws.append([
        "name", "duty_type_name", "duty_location_name", "recurrence_type", "weekdays",
        "start_time", "end_time", "required_count", "auto_roll", "auto_roll_until",
        "duration_days", "notes", "eligible_units",
    ])
    for r in rows:
        ws.append(r)
    return wb


def test_shift_template_resolves_duty_type_and_location(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", dt.name, loc.name, "weekdays", "", "08:00", "17:00", 1, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["shift_templates"][0]
    assert row["action"] == "new"
    assert row["resolved_duty_type_id"] == str(dt.id)
    assert row["resolved_duty_location_id"] == str(loc.id)


def test_shift_template_unresolved_duty_type_errors(admin_session):
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", "לא קיים", loc.name, "weekdays", "", "", "", 1, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["shift_templates"][0]
    assert row["action"] == "error"
    assert any("סוג תורנות לא מזוהה" in e for e in row["errors"])


def test_shift_template_matches_existing_by_name_for_update(admin_session):
    from app.services.shift_templates import create_template

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    tpl_name = f"tpl_{_uid()}"
    create_template(
        admin_session, name=tpl_name, duty_type_id=dt.id, duty_location_id=loc.id,
        recurrence_type="weekdays", weekdays=[],
    )
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [tpl_name, dt.name, loc.name, "weekdays", "", "", "", 2, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["shift_templates"][0]
    assert row["action"] == "update"
    assert row["existing_id"] is not None


def test_shift_template_eligible_unit_resolution(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", dt.name, loc.name, "weekdays", "", "", "", 1, "false", "", 1, "", node.name],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row = sess.parsed_state["shift_templates"][0]
    assert row["resolved_eligible_node_ids"] == [str(node.id)]


def test_shift_template_field_override_changes_resolved_name(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", dt.name, loc.name, "weekdays", "", "", "", 1, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["shift_templates"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"shift_templates": {str(row_num): {"required_count": 5}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["shift_templates"][0]
    assert row["required_count"] == 5


def test_shift_template_field_override_bypasses_eligible_unit_resolution(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", dt.name, loc.name, "weekdays", "", "", "", 1, "false", "", 1, "", "לא קיים"],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["shift_templates"][0]["row"]
    # Without the override, the unresolvable "לא קיים" eligible unit errors the row.
    assert sess.parsed_state["shift_templates"][0]["action"] == "error"

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"shift_templates": {str(row_num): {"resolved_eligible_node_ids": []}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["shift_templates"][0]
    assert row["action"] == "new"
    assert row["resolved_eligible_node_ids"] == []


def test_shift_template_field_override_none_value_still_bypasses_resolution(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [f"tpl_{_uid()}", dt.name, loc.name, "weekdays", "", "", "", 1, "false", "", 1, "", "לא קיים"],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["shift_templates"][0]["row"]
    assert sess.parsed_state["shift_templates"][0]["action"] == "error"

    # Override value is explicitly None (key present, value None) — this is the input that
    # discriminates key-presence detection (correctly skips resolution) from value-identity
    # detection (would incorrectly re-run resolution, since a plain `is None` check on the
    # resolved value can't tell "override explicitly set to None" apart from "no override at
    # all", which also yields None).
    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"shift_templates": {str(row_num): {"resolved_eligible_node_ids": None}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["shift_templates"][0]
    assert row["action"] == "new"
    assert row["resolved_eligible_node_ids"] == []


def test_confirm_session_creates_shift_template(admin_session):
    from app.db.models import ShiftTemplate as ShiftTemplateModel

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    tpl_name = f"tpl_{_uid()}"
    wb = _wb_with_shift_templates([
        [tpl_name, dt.name, loc.name, "weekdays", "", "08:00", "17:00", 2, "false", "", 1, "note", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["created"] == 1
    assert len(sess.created_links["shift_templates"]) == 1
    created = admin_session.get(ShiftTemplateModel, uuid.UUID(sess.created_links["shift_templates"][0]))
    assert created is not None
    assert created.name == tpl_name
    assert created.duty_type_id == dt.id
    assert created.required_count == 2
    assert created.notes == "note"


def test_confirm_session_updates_existing_shift_template(admin_session):
    from app.services.shift_templates import create_template

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    tpl_name = f"tpl_{_uid()}"
    tpl = create_template(
        admin_session, name=tpl_name, duty_type_id=dt.id, duty_location_id=loc.id,
        recurrence_type="weekdays", weekdays=[], required_count=1,
    )
    admin_session.commit()

    wb = _wb_with_shift_templates([
        [tpl_name, dt.name, loc.name, "weekdays", "", "", "", 3, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["updated"] == 1
    admin_session.refresh(tpl)
    assert tpl.required_count == 3


def test_confirm_session_shift_template_update_preserves_unspecified_fields(admin_session):
    from app.services.shift_templates import create_template

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    tpl_name = f"tpl_{_uid()}"
    tpl = create_template(
        admin_session, name=tpl_name, duty_type_id=dt.id, duty_location_id=loc.id,
        recurrence_type="weekdays", weekdays=[], required_count=1,
        notes="existing note", eligible_node_ids=[node.id],
    )
    admin_session.commit()

    # Re-import the same template with notes/eligible_units left blank in the sheet,
    # only changing required_count.
    wb = _wb_with_shift_templates([
        [tpl_name, dt.name, loc.name, "weekdays", "", "", "", 5, "false", "", 1, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["updated"] == 1
    admin_session.refresh(tpl)
    assert tpl.required_count == 5
    assert tpl.notes == "existing note"  # not cleared
    assert tpl.eligible_node_ids == [node.id]  # not cleared


def test_confirm_session_shift_template_update_preserves_blank_recurrence_fields(admin_session):
    from app.services.shift_templates import create_template

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    tpl_name = f"tpl_{_uid()}"
    # recurrence_type must be "weekly" here: update_template forces duration_days
    # back to 1 for any non-"weekly" recurrence (a separate, deliberate rule in
    # update_template itself, unrelated to this blank-cell bug) — so only
    # "weekly" lets duration_days > 1 survive an update to exercise the fix.
    # weekdays is filled in the re-import row below (not left blank), since
    # blank-cell preservation for `weekdays` itself is not part of this fix.
    tpl = create_template(
        admin_session, name=tpl_name, duty_type_id=dt.id, duty_location_id=loc.id,
        recurrence_type="weekly", weekdays=[1, 2, 3], required_count=7, duration_days=3, auto_roll=True,
        auto_roll_until=None,
    )
    admin_session.commit()

    # Re-import leaving recurrence_type/required_count/duration_days/auto_roll blank,
    # only specifying notes (and weekdays, kept identical to the existing template).
    # auto_roll is left as a genuinely-empty cell (None) rather than "" — parse_bool
    # only treats a true None cell as "blank" (an empty *string* cell parses to
    # False, same as a real openpyxl blank cell always yields None, never "").
    wb = _wb_with_shift_templates([
        [tpl_name, dt.name, loc.name, "", "1,2,3", "", "", "", None, "", "", "updated note", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    admin_session.commit()

    result = confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    assert result["updated"] == 1
    admin_session.refresh(tpl)
    assert tpl.notes == "updated note"
    assert tpl.recurrence_type == "weekly"  # preserved, not reset to "weekdays"
    assert tpl.required_count == 7  # preserved, not reset to 1
    assert tpl.duration_days == 3  # preserved, not reset to 1
    assert tpl.auto_roll is True  # preserved, not reset to False


def test_hierarchy_field_override_changes_level(admin_session):
    wb = _wb_with_hierarchy([
        [f"node_{_uid()}", "branch", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["hierarchy"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"hierarchy": {str(row_num): {"level": "team"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["hierarchy"][0]
    assert row["level"] == "team"


def test_soldier_field_override_changes_rank(admin_session):
    wb = _wb_with_soldiers([
        ["1234567", "Some Soldier", "rank1", "", "", "", "", "", "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["soldiers"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"soldiers": {str(row_num): {"rank": "rank2"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["soldiers"][0]
    assert row["rank"] == "rank2"


def test_duty_shift_field_override_changes_required_count(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["duty_shifts"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"duty_shifts": {str(row_num): {"required_count": 9}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["duty_shifts"][0]
    assert row["required_count"] == 9


def test_assignment_field_override_changes_notes(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sld_{_uid()}")
    admin_session.commit()

    from app.db.models import DutyShift
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=5,
    )
    admin_session.add(shift)
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["assignments"][0]["row"]

    set_selections(admin_session, session_id=sess.id, selections={
        "_field_overrides": {"assignments": {str(row_num): {"notes": "overridden note"}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["assignments"][0]
    assert row["notes"] == "overridden note"


def test_assignment_duty_type_resolved_via_by_row_mapping(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sld_{_uid()}")
    admin_session.commit()

    from app.db.models import DutyShift
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        required_count=5,
    )
    admin_session.add(shift)
    admin_session.commit()

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, "no_such_duty_type", loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    sess = create_session(
        admin_session, filename="f.xlsx", content=_to_bytes(wb), actor=admin, parser_id="v1_standard",
    )
    row_num = sess.parsed_state["assignments"][0]["row"]
    assert sess.parsed_state["assignments"][0]["action"] == "error"

    set_selections(admin_session, session_id=sess.id, selections={
        "_name_mappings": {"duty_type": {"by_row": {f"assignments:{row_num}": str(dt.id)}}},
    })
    admin_session.commit()

    reparse_session(admin_session, session_id=sess.id, actor=admin)
    row = sess.parsed_state["assignments"][0]
    assert row["action"] == "new"
