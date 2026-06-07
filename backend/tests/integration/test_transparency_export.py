import io
import uuid

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from tests.helpers import auth_headers, create_node, create_soldier


def test_soldiers_export_returns_xlsx(client: TestClient, admin_session: Session):
    s = create_soldier(admin_session, personal_number="5700001", role="soldier")
    r = client.get("/api/scoring/transparency/export", headers=auth_headers(s))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "חיילים" in wb.sheetnames


def test_soldiers_export_contains_header_row(client: TestClient, admin_session: Session):
    s = create_soldier(admin_session, personal_number="5700002", role="soldier")
    r = client.get("/api/scoring/transparency/export", headers=auth_headers(s))
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["חיילים"]
    headers = [ws.cell(1, col).value for col in range(1, 11)]
    assert headers == [
        "יחידה / תת-יחידה", "שם", "יחידה", "תאריך הצטרפות", "ימים פעילים", "דרגה",
        "כמות משמרות", "ניקוד מצטבר", "ניקוד ליום", "ניקוד מנורמל",
    ]


def test_soldiers_export_node_filter(client: TestClient, admin_session: Session):
    admin = create_soldier(admin_session, personal_number="5700003", role="admin")
    root = create_node(admin_session, level="division", name="root-exp")
    child = create_node(admin_session, level="unit", name="child-exp", parent=root)
    s_in = create_soldier(
        admin_session, personal_number="5700004", role="soldier",
        hierarchy_node_id=child.id,
    )
    s_out = create_soldier(
        admin_session, personal_number="5700005", role="soldier",
        hierarchy_node_id=root.id,
    )
    # filter by child node — only s_in should appear (root is not in child's subtree)
    r = client.get(
        f"/api/scoring/transparency/export?node_id={child.id}",
        headers=auth_headers(admin),
    )
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["חיילים"]
    # column 1 is node path, column 2 is soldier full_name
    names = [ws.cell(row, 2).value for row in range(2, ws.max_row + 1)]
    assert s_in.full_name in names
    assert s_out.full_name not in names


def test_soldiers_export_unknown_node_returns_404(client: TestClient, admin_session: Session):
    s = create_soldier(admin_session, personal_number="5700006", role="soldier")
    r = client.get(
        f"/api/scoring/transparency/export?node_id={uuid.uuid4()}",
        headers=auth_headers(s),
    )
    assert r.status_code == 404
    assert r.json()["detail"] == "not_found"


def test_soldiers_export_requires_auth(client: TestClient):
    r = client.get("/api/scoring/transparency/export")
    assert r.status_code == 401


def test_sub_units_export_returns_xlsx(client: TestClient, admin_session: Session):
    s = create_soldier(admin_session, personal_number="5700010", role="soldier")
    r = client.get("/api/scoring/transparency/sub-units/export", headers=auth_headers(s))
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "תתי יחידות" in wb.sheetnames


def test_sub_units_export_contains_header_row(client: TestClient, admin_session: Session):
    s = create_soldier(admin_session, personal_number="5700011", role="soldier")
    r = client.get("/api/scoring/transparency/sub-units/export", headers=auth_headers(s))
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["תתי יחידות"]
    headers = [ws.cell(1, col).value for col in range(1, 9)]
    assert headers == [
        "יחידה", "כמות חיילים", "חיילים פעילים (%)",
        "ממוצע ימים פעילים", "ממוצע ניקוד לחייל",
        "ממוצע ניקוד לחייל פעיל", "ניקוד ליום (מסגרת)", "ניקוד מנורמל ממוצע",
    ]


def test_sub_units_export_aggregates_per_node(client: TestClient, admin_session: Session):
    admin = create_soldier(admin_session, personal_number="5700012", role="admin")
    root = create_node(admin_session, level="division", name="root-su-exp")
    child = create_node(admin_session, level="unit", name="child-su-exp", parent=root)
    create_soldier(admin_session, personal_number="5700013", role="soldier",
                   hierarchy_node_id=child.id)
    create_soldier(admin_session, personal_number="5700014", role="soldier",
                   hierarchy_node_id=child.id)
    r = client.get("/api/scoring/transparency/sub-units/export", headers=auth_headers(admin))
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["תתי יחידות"]
    node_names = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    # Both root and child appear (child has 2 soldiers; root has them via path)
    assert "child-su-exp" in node_names
    assert "root-su-exp" in node_names
    # child row should show count == 2
    child_row_idx = next(i for i, n in enumerate(node_names, start=2) if n == "child-su-exp")
    assert ws.cell(child_row_idx, 2).value == 2


def test_sub_units_export_requires_auth(client: TestClient):
    r = client.get("/api/scoring/transparency/sub-units/export")
    assert r.status_code == 401
