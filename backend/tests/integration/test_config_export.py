from __future__ import annotations

import io
import uuid
from decimal import Decimal

import openpyxl

from app.db.models import DutyLocation
from app.services.duty_config import create_duty_type, create_exemption_type, set_exemption_duty_types
from app.services.hierarchy import create_node, set_commander
from tests.helpers import auth_headers, create_range_location, create_soldier


def _uid() -> str:
    return uuid.uuid4().hex[:8]


def _token(soldier) -> str:
    return auth_headers(soldier)["Authorization"].split(" ", 1)[1]


def test_export_returns_only_requested_sheets(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    loc = DutyLocation(name=f"שער_{_uid()}")
    admin_session.add(loc)
    admin_session.commit()

    resp = client.get(
        "/api/config/export?sheets=duty_locations",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["מיקומי תפקיד"]
    rows = list(wb["מיקומי תפקיד"].iter_rows(min_row=2, values_only=True))
    assert any(r[0] == loc.name for r in rows)


def test_export_defaults_to_all_sheets(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = client.get(
        "/api/config/export", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {
        "סוגי תפקידים",
        "מיקומי תפקיד",
        "היררכיה",
        "סוגי פטורים",
        "הגדרות מערכת",
        "דיווחי באגים",
        "מיקומי מטווח",
    }


def test_export_includes_range_locations(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    loc = create_range_location(admin_session, name=f"מטווח_{_uid()}")
    admin_session.commit()

    resp = client.get(
        "/api/config/export?sheets=range_locations",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["מיקומי מטווח"]
    rows = list(wb["מיקומי מטווח"].iter_rows(values_only=True))
    assert rows[0] == ("שם", "פעיל")
    assert any(r[0] == loc.name for r in rows[1:])


def test_export_hierarchy_includes_commander_and_duty_managers(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    commander = create_soldier(admin_session, personal_number=f"cmd_{_uid()}")
    node = create_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    set_commander(admin_session, node_id=node.id, commander_id=commander.id, actor_id=admin.id)
    admin_session.commit()

    resp = client.get(
        "/api/config/export?sheets=hierarchy", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["היררכיה"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == node.name)
    assert row[3] == commander.personal_number  # commander_personal_number column


def test_export_exemption_types_includes_applies_to(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    et = create_exemption_type(admin_session, name=f"et_{_uid()}")
    set_exemption_duty_types(admin_session, exemption_type_id=et.id, duty_type_ids=[dt.id])
    admin_session.commit()

    resp = client.get(
        "/api/config/export?sheets=exemption_types", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["סוגי פטורים"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == et.name)
    assert dt.name in row[-1]


def test_export_duty_type_with_no_requirements_has_blank_json_cell(client, admin_session):
    # Regression guard: a duty type with no custom requirements must export
    # a blank requirements_json cell, not the literal string "null" — a
    # human hand-editing the sheet shouldn't see stray "null" text, and a
    # blank cell is what the import parser treats as "no requirements".
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    admin_session.commit()

    resp = client.get(
        "/api/config/export?sheets=duty_types", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["סוגי תפקידים"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == dt.name)
    assert row[-1] in (None, "")
