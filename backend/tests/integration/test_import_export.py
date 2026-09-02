from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest

from sqlalchemy import select

from app.db.models import DutyAssignment, DutyLocation, DutyShift, DutyShiftNodeQuota, RangeAssignment, RangeEvent, TelegramLink
from app.services.duty_config import create_duty_type
from app.services.excel_bilingual import HE_HEADERS
from app.services.import_sessions import confirm_session, create_session
from tests.helpers import (
    auth_headers,
    create_node,
    create_range_assignment,
    create_range_event,
    create_range_location,
    create_soldier,
)


def _uid() -> str:
    import uuid
    return uuid.uuid4().hex[:8]


def test_export_round_trips_soldiers_duty_shifts_and_assignments(client, admin_session):
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}", hierarchy_node_id=node.id)
    soldier.unit_join_date = date(2024, 1, 15)
    soldier.food_type = "vegetarian"
    soldier.food_constraints = "No dairy"
    soldier.rank = "טוראי"
    soldier.rank_track = "enlisted"
    soldier.next_rank_date = date(2027, 1, 15)
    soldier.next_rank_date_overridden = True
    soldier.current_rank_since = date(2025, 1, 15)
    soldier.profile_picture_url = "https://example.test/profile.png"
    admin_session.add(TelegramLink(
        soldier_id=soldier.id,
        telegram_chat_id=123456789,
        telegram_username="soldier_test",
        is_verified=True,
        notifications_enabled=False,
        verified_at=datetime(2026, 1, 15, 10, 30),
    ))
    shift = DutyShift(
        duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date(2024, 6, 15), end_date=date(2024, 6, 16),
        required_count=2,
    )
    admin_session.add(shift)
    admin_session.flush()
    admin_session.add(DutyShiftNodeQuota(duty_shift_id=shift.id, hierarchy_node_id=node.id, count=1))
    admin_session.add(DutyAssignment(
        soldier_id=soldier.id, duty_type_id=dt.id, duty_location_id=loc.id,
        duty_shift_id=shift.id, start_date=shift.start_date, end_date=shift.end_date,
    ))
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.commit()

    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/export", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {
        "חיילים", "משמרות", "שיבוצים", "תבניות משמרות",
        "ימי מטווח", "שיבוצי מטווח", "מועדי קידום",
    }

    soldiers_sheet = wb["חיילים"]
    soldier_headers = [cell.value for cell in next(soldiers_sheet.iter_rows(min_row=1, max_row=1))]
    soldier_header_index = {}
    for index, header in enumerate(soldier_headers):
        soldier_header_index[header] = index
        soldier_header_index[HE_HEADERS.get(header, header)] = index
        for english, hebrew in HE_HEADERS.items():
            if header == hebrew:
                soldier_header_index[english] = index
    soldier_rows = list(soldiers_sheet.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == soldier.personal_number for r in soldier_rows)
    soldier_row = next(r for r in soldier_rows if r[0] == soldier.personal_number)
    assert soldier_row[9] == "15.01.2024"  # unit_join_date
    assert soldier_row[13] == "vegetarian"  # food_type
    assert soldier_row[14] == "No dairy"  # food_constraints
    assert soldier_row[soldier_header_index["profile_picture_url"]] == "https://example.test/profile.png"
    assert soldier_row[soldier_header_index["telegram_chat_id"]] == 123456789
    assert soldier_row[soldier_header_index["telegram_username"]] == "soldier_test"
    assert soldier_row[soldier_header_index["telegram_is_verified"]] == "true"
    assert soldier_row[soldier_header_index["telegram_notifications_enabled"]] == "false"
    assert soldier_row[soldier_header_index["telegram_verified_at"]] == "2026-01-15T10:30:00+00:00"
    assert soldier_row[soldier_header_index["next_rank_date_overridden"]] == "true"
    assert "current_rank_since" in soldier_header_index or HE_HEADERS["current_rank_since"] in soldier_header_index, soldier_headers
    assert soldier_row[soldier_header_index.get("current_rank_since", soldier_header_index[HE_HEADERS["current_rank_since"]])] == "15.01.2025"

    import_session = create_session(
        admin_session,
        filename="soldiers-roundtrip.xlsx",
        content=resp.content,
        actor=admin,
        parser_id="v1_standard",
    )
    parsed_soldier = next(
        r for r in import_session.parsed_state["soldiers"]
        if r["personal_number"] == soldier.personal_number
    )
    assert parsed_soldier["unit_join_date"] == "2024-01-15"
    assert parsed_soldier["food_type"] == "vegetarian"
    assert parsed_soldier["food_constraints"] == "No dairy"
    assert parsed_soldier["profile_picture_url"] == "https://example.test/profile.png"
    assert parsed_soldier["telegram_chat_id"] == 123456789
    assert parsed_soldier["telegram_username"] == "soldier_test"
    assert parsed_soldier["telegram_is_verified"] is True
    assert parsed_soldier["telegram_notifications_enabled"] is False
    assert parsed_soldier["telegram_verified_at"] == "2026-01-15T10:30:00+00:00"
    assert parsed_soldier["next_rank_date"] == "2027-01-15"
    assert parsed_soldier["next_rank_date_overridden"] is True
    assert parsed_soldier["current_rank_since"] == "2025-01-15"
    assert parsed_soldier["action"] == "update"
    confirm_session(admin_session, session_id=import_session.id, actor=admin)
    admin_session.refresh(soldier)
    assert soldier.unit_join_date == date(2024, 1, 15)
    assert soldier.food_type == "vegetarian"
    assert soldier.food_constraints == "No dairy"
    assert soldier.profile_picture_url == "https://example.test/profile.png"
    link = admin_session.execute(
        select(TelegramLink).where(TelegramLink.soldier_id == soldier.id)
    ).scalar_one()
    assert link.telegram_chat_id == 123456789
    assert link.telegram_username == "soldier_test"
    assert link.is_verified is True
    assert link.notifications_enabled is False
    assert link.verified_at.isoformat() == "2026-01-15T10:30:00+00:00"
    assert soldier.rank == "טוראי"
    assert soldier.rank_track == "enlisted"
    assert soldier.next_rank_date == date(2027, 1, 15)
    assert soldier.next_rank_date_overridden is True
    assert soldier.current_rank_since == date(2025, 1, 15)

    shift_rows = list(wb["משמרות"].iter_rows(min_row=2, values_only=True))
    matching_shift = next(r for r in shift_rows if r[0] == dt.name and r[1] == loc.name)
    assert matching_shift[6] == 2  # required_count
    assert node.name in matching_shift[7]  # node_quotas string

    assignment_rows = list(wb["שיבוצים"].iter_rows(min_row=2, values_only=True))
    assert len(assignment_rows) == 1
    a = assignment_rows[0]
    assert a[0] == soldier.personal_number
    assert a[1] == soldier.full_name
    assert a[2] == dt.name
    assert a[3] == loc.name


def test_export_omits_assignments_without_linked_shift(client, admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.add(DutyAssignment(
        soldier_id=soldier.id, duty_type_id=dt.id, duty_location_id=loc.id,
        start_date=date(2024, 6, 15), end_date=date(2024, 6, 16),
    ))  # no duty_shift_id
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.commit()

    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/export", headers={"Authorization": f"Bearer {token}"})
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assignment_rows = list(wb["שיבוצים"].iter_rows(min_row=2, values_only=True))
    assert not any(r[0] == soldier.personal_number for r in assignment_rows)


def test_import_export_includes_range_events_and_assignments(client, admin_session):
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    loc = create_range_location(admin_session, name=f"מטווח_{_uid()}")
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}", hierarchy_node_id=node.id)
    event = create_range_event(
        admin_session, hierarchy_node=node, range_location=loc,
        range_type="live", event_date=date(2024, 6, 20),
    )
    create_range_assignment(admin_session, range_event=event, soldier=soldier)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.commit()

    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get(
        "/api/import/export?sheets=range_events,range_assignments",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {"ימי מטווח", "שיבוצי מטווח"}

    event_rows = list(wb["ימי מטווח"].iter_rows(min_row=2, values_only=True))
    assert any(r[0] == node.name and r[3] == loc.name for r in event_rows)

    assignment_rows = list(wb["שיבוצי מטווח"].iter_rows(min_row=2, values_only=True))
    assert any(r[0] == soldier.personal_number for r in assignment_rows)


def test_range_events_and_assignments_export_import_round_trip(client, admin_session):
    """Genuine export -> re-upload -> confirm -> verify-DB round trip for
    import_export's range_events/range_assignments sheets (Task 10), per
    Finding 4. This is the regression guard that would have caught the
    Enum-serialization bug already fixed in Task 10 (RangeType/
    RangeEventStatus/RangeAttendanceStatus round-tripping through the xlsx
    cell as their real string value, not a Python repr like
    "RangeType.live") — a bad serialization would surface here as either an
    exception during export, or an "invalid range_type/status" row error on
    re-import."""
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    loc = create_range_location(admin_session, name=f"מטווח_{_uid()}")
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}", hierarchy_node_id=node.id)
    manager = create_soldier(admin_session, personal_number=f"dm_{_uid()}", role="duty_manager")
    event = create_range_event(
        admin_session, hierarchy_node=node, range_location=loc,
        range_type="live", event_date=date(2024, 6, 20), required_count=3,
    )
    event.responsible_duty_manager_id = manager.id
    create_range_assignment(admin_session, range_event=event, soldier=soldier, attendance_status="present")
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.commit()

    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get(
        "/api/import/export?sheets=range_events,range_assignments",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200

    sess = create_session(
        admin_session, filename="roundtrip.xlsx", content=resp.content, actor=admin, parser_id="v1_standard",
    )
    # The export contains every range_event/range_assignment in the DB (not
    # just this test's), so match on this test's unique node/location/soldier
    # names rather than assuming index 0 — other tests sharing this worker's
    # DB may have already created their own rows.
    event_row = next(
        r for r in sess.parsed_state["range_events"]
        if r["hierarchy_node_name"] == node.name and r["range_location_name"] == loc.name
    )
    assignment_row = next(
        r for r in sess.parsed_state["range_assignments"] if r["personal_number"] == soldier.personal_number
    )
    # No "update" path exists for range_events (a fresh import always proposes
    # "new" — see _resolve_range_events), but the re-parsed enum/date/location
    # fields must be valid and error-free; the assignment row must resolve
    # against the *existing* DB event (not the about-to-be-recreated one) and
    # therefore detect it's already assigned.
    assert event_row["errors"] == []
    assert event_row["action"] == "new"
    assert event_row["range_type"] == "live"
    assert event_row["status"] == "planned"
    assert event_row["responsible_duty_manager_personal_number"] == manager.personal_number
    assert event_row["responsible_duty_manager_id"] == str(manager.id)
    assert assignment_row["errors"] == []
    assert assignment_row["resolved_range_event_id"] == str(event.id)
    assert assignment_row["action"] == "skip"  # soldier already assigned to this exact event

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    events = admin_session.execute(
        select(RangeEvent).where(RangeEvent.hierarchy_node_id == node.id, RangeEvent.range_location_id == loc.id)
    ).scalars().all()
    assert len(events) == 2  # the original + the re-imported "new" row
    assert any(e.required_count == 3 and e.range_type == "live" for e in events)
    assert any(
        e.required_count == 3
        and e.range_type == "live"
        and e.responsible_duty_manager_id == manager.id
        for e in events
    )

    assignments = admin_session.execute(
        select(RangeAssignment).where(RangeAssignment.soldier_id == soldier.id)
    ).scalars().all()
    assert len(assignments) == 1  # the "skip" action prevented a duplicate assignment
    assert assignments[0].attendance_status.value == "present"
