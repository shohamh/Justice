from __future__ import annotations

import io
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import select

from app.db.models import DutyLocation
from app.services.duty_config import create_duty_type
import app.services.import_parsers.v1_standard  # noqa: F401  (registers "v1_standard" parser)
from tests.helpers import auth_headers, create_node, create_soldier


def _uid(n=1):
    import uuid
    return uuid.uuid4().hex[:8]


def _wb_with_duty_shifts(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("duty_shifts")
    ws.append([
        "duty_type_name", "duty_location_name", "start_date", "end_date",
        "start_time", "end_time", "required_count", "node_quotas", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _wb_with_assignments(rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("assignments")
    ws.append([
        "personal_number", "full_name", "duty_type_name", "duty_location_name",
        "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes",
    ])
    for r in rows:
        ws.append(r)
    return wb


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _token(soldier) -> str:
    return auth_headers(soldier)["Authorization"].split(" ", 1)[1]


def _upload(client, token, xlsx: bytes, parser_id="v1_standard"):
    return client.post(
        f"/api/import/sessions?parser_id={parser_id}",
        files={"file": ("import.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


def _make_wb_bytes(admin_session):
    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()
    wb = _wb_with_duty_shifts([
        [dt.name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    return _to_bytes(wb), dt, loc


def test_upload_creates_draft_session(client, admin_session):
    xlsx, dt, loc = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = _upload(client, _token(admin), xlsx)

    assert resp.status_code == 200
    body = resp.json()
    assert "session_id" in body
    assert body["preview"]["duty_shifts"][0]["action"] == "new"


def test_upload_invalid_file_type_400(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = client.post(
        "/api/import/sessions",
        files={"file": ("import.xlsx", b"not an xlsx", "application/octet-stream")},
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 400


def test_upload_wrong_extension_400(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = client.post(
        "/api/import/sessions",
        files={"file": ("import.docx", xlsx, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")},
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 400
    assert resp.json()["detail"] == "invalid_file_type"


def test_list_sessions_respects_status_filter_and_ownership(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    dm = create_soldier(admin_session, personal_number=f"dm_{_uid()}", role="duty_manager", hierarchy_node_id=node.id)

    resp1 = _upload(client, _token(admin), xlsx)
    assert resp1.status_code == 200

    xlsx2, _, _ = _make_wb_bytes(admin_session)
    resp2 = _upload(client, _token(dm), xlsx2)
    assert resp2.status_code == 200

    # dm sees only their own session
    dm_list = client.get(
        "/api/import/sessions", headers={"Authorization": f"Bearer {_token(dm)}"}
    ).json()
    dm_ids = {s["id"] for s in dm_list}
    assert resp2.json()["session_id"] in dm_ids
    assert resp1.json()["session_id"] not in dm_ids

    # admin sees all
    admin_list = client.get(
        "/api/import/sessions", headers={"Authorization": f"Bearer {_token(admin)}"}
    ).json()
    admin_ids = {s["id"] for s in admin_list}
    assert resp1.json()["session_id"] in admin_ids
    assert resp2.json()["session_id"] in admin_ids

    # status filter
    filtered = client.get(
        "/api/import/sessions?status_filter=cancelled",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    ).json()
    assert resp1.json()["session_id"] not in {s["id"] for s in filtered}


def test_get_detail_404_and_403(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    dm = create_soldier(admin_session, personal_number=f"dm_{_uid()}", role="duty_manager", hierarchy_node_id=node.id)
    other_node = create_node(admin_session, level="branch", name=f"node2_{_uid()}")
    other_dm = create_soldier(
        admin_session, personal_number=f"dm2_{_uid()}", role="duty_manager", hierarchy_node_id=other_node.id
    )

    resp = _upload(client, _token(dm), xlsx)
    session_id = resp.json()["session_id"]

    not_found = client.get(
        "/api/import/sessions/00000000-0000-0000-0000-000000000000",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert not_found.status_code == 404

    forbidden = client.get(
        f"/api/import/sessions/{session_id}",
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    ok = client.get(
        f"/api/import/sessions/{session_id}",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert ok.status_code == 200
    assert ok.json()["id"] == session_id


def test_reparse_flips_row_after_duty_type_created(client, admin_session):
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    admin_session.commit()

    dt_name = f"dt_{_uid()}"
    wb = _wb_with_duty_shifts([
        [dt_name, loc.name, "15.06.2024", "16.06.2024", "", "", 5, "", ""],
    ])
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    resp = _upload(client, _token(admin), _to_bytes(wb))
    session_id = resp.json()["session_id"]
    assert resp.json()["preview"]["duty_shifts"][0]["action"] == "error"

    create_duty_type(admin_session, name=dt_name, score_per_day=Decimal("1.00"))
    admin_session.commit()

    reparsed = client.post(
        f"/api/import/sessions/{session_id}/reparse",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert reparsed.status_code == 200
    assert reparsed.json()["parsed_state"]["duty_shifts"][0]["action"] == "new"


def test_confirm_applies_and_returns_counts(client, admin_session):
    xlsx, dt, loc = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = _upload(client, _token(admin), xlsx)
    session_id = resp.json()["session_id"]

    confirmed = client.post(
        f"/api/import/sessions/{session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert confirmed.status_code == 200
    body = confirmed.json()
    assert body["created"] == 1
    assert body["skipped"] == 0
    assert body["errors"] == []


def test_cancel_sets_status(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = _upload(client, _token(admin), xlsx)
    session_id = resp.json()["session_id"]

    cancelled = client.post(
        f"/api/import/sessions/{session_id}/cancel",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "cancelled"


def test_done_requires_confirmed_first(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = _upload(client, _token(admin), xlsx)
    session_id = resp.json()["session_id"]

    too_early = client.post(
        f"/api/import/sessions/{session_id}/done",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert too_early.status_code == 400

    confirmed = client.post(
        f"/api/import/sessions/{session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert confirmed.status_code == 200

    done = client.post(
        f"/api/import/sessions/{session_id}/done",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert done.status_code == 200
    assert done.json()["status"] == "done"


def test_selections_patch_returns_ok(client, admin_session):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = _upload(client, _token(admin), xlsx)
    session_id = resp.json()["session_id"]

    patched = client.patch(
        f"/api/import/sessions/{session_id}/selections",
        json={"selections": {"duty_shifts": {"2": "skip"}}},
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert patched.status_code == 200
    assert patched.json() == {"ok": True}


def _dm_and_other_dm(admin_session):
    node = create_node(admin_session, level="branch", name=f"node_{_uid()}")
    dm = create_soldier(admin_session, personal_number=f"dm_{_uid()}", role="duty_manager", hierarchy_node_id=node.id)
    other_node = create_node(admin_session, level="branch", name=f"node2_{_uid()}")
    other_dm = create_soldier(
        admin_session, personal_number=f"dm2_{_uid()}", role="duty_manager", hierarchy_node_id=other_node.id
    )
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    return dm, other_dm, admin


def _upload_owned_session(client, admin_session, owner):
    xlsx, _, _ = _make_wb_bytes(admin_session)
    resp = _upload(client, _token(owner), xlsx)
    assert resp.status_code == 200
    return resp.json()["session_id"]


def test_reparse_enforces_ownership(client, admin_session):
    dm, other_dm, admin = _dm_and_other_dm(admin_session)

    forbidden_session_id = _upload_owned_session(client, admin_session, dm)
    forbidden = client.post(
        f"/api/import/sessions/{forbidden_session_id}/reparse",
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    admin_ok_session_id = _upload_owned_session(client, admin_session, dm)
    admin_ok = client.post(
        f"/api/import/sessions/{admin_ok_session_id}/reparse",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert admin_ok.status_code == 200

    owner_ok_session_id = _upload_owned_session(client, admin_session, dm)
    owner_ok = client.post(
        f"/api/import/sessions/{owner_ok_session_id}/reparse",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert owner_ok.status_code == 200


def test_selections_enforces_ownership(client, admin_session):
    dm, other_dm, admin = _dm_and_other_dm(admin_session)

    forbidden_session_id = _upload_owned_session(client, admin_session, dm)
    forbidden = client.patch(
        f"/api/import/sessions/{forbidden_session_id}/selections",
        json={"selections": {"duty_shifts": {"2": "skip"}}},
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    admin_ok_session_id = _upload_owned_session(client, admin_session, dm)
    admin_ok = client.patch(
        f"/api/import/sessions/{admin_ok_session_id}/selections",
        json={"selections": {"duty_shifts": {"2": "skip"}}},
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert admin_ok.status_code == 200

    owner_ok_session_id = _upload_owned_session(client, admin_session, dm)
    owner_ok = client.patch(
        f"/api/import/sessions/{owner_ok_session_id}/selections",
        json={"selections": {"duty_shifts": {"2": "skip"}}},
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert owner_ok.status_code == 200


def test_confirm_enforces_ownership(client, admin_session):
    dm, other_dm, admin = _dm_and_other_dm(admin_session)

    forbidden_session_id = _upload_owned_session(client, admin_session, dm)
    forbidden = client.post(
        f"/api/import/sessions/{forbidden_session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    admin_ok_session_id = _upload_owned_session(client, admin_session, dm)
    admin_ok = client.post(
        f"/api/import/sessions/{admin_ok_session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert admin_ok.status_code == 200

    owner_ok_session_id = _upload_owned_session(client, admin_session, dm)
    owner_ok = client.post(
        f"/api/import/sessions/{owner_ok_session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert owner_ok.status_code == 200


def test_cancel_enforces_ownership(client, admin_session):
    dm, other_dm, admin = _dm_and_other_dm(admin_session)

    forbidden_session_id = _upload_owned_session(client, admin_session, dm)
    forbidden = client.post(
        f"/api/import/sessions/{forbidden_session_id}/cancel",
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    admin_ok_session_id = _upload_owned_session(client, admin_session, dm)
    admin_ok = client.post(
        f"/api/import/sessions/{admin_ok_session_id}/cancel",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert admin_ok.status_code == 200

    owner_ok_session_id = _upload_owned_session(client, admin_session, dm)
    owner_ok = client.post(
        f"/api/import/sessions/{owner_ok_session_id}/cancel",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert owner_ok.status_code == 200


def test_done_enforces_ownership(client, admin_session):
    dm, other_dm, admin = _dm_and_other_dm(admin_session)

    forbidden_session_id = _upload_owned_session(client, admin_session, dm)
    forbidden = client.post(
        f"/api/import/sessions/{forbidden_session_id}/done",
        headers={"Authorization": f"Bearer {_token(other_dm)}"},
    )
    assert forbidden.status_code == 403

    admin_ok_session_id = _upload_owned_session(client, admin_session, dm)
    confirm_admin = client.post(
        f"/api/import/sessions/{admin_ok_session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert confirm_admin.status_code == 200
    admin_ok = client.post(
        f"/api/import/sessions/{admin_ok_session_id}/done",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert admin_ok.status_code == 200

    owner_ok_session_id = _upload_owned_session(client, admin_session, dm)
    confirm_owner = client.post(
        f"/api/import/sessions/{owner_ok_session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert confirm_owner.status_code == 200
    owner_ok = client.post(
        f"/api/import/sessions/{owner_ok_session_id}/done",
        headers={"Authorization": f"Bearer {_token(dm)}"},
    )
    assert owner_ok.status_code == 200


def test_upload_and_confirm_assignments_end_to_end(client, admin_session):
    from app.db.models import DutyAssignment

    dt = create_duty_type(admin_session, name=f"dt_{_uid()}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add(loc)
    admin_session.flush()
    soldier = create_soldier(admin_session, personal_number=f"sol_{_uid()}")
    admin_session.commit()
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")

    wb = _wb_with_assignments([
        [soldier.personal_number, soldier.full_name, dt.name, loc.name,
         "15.06.2024", "16.06.2024", "", "", "false", ""],
    ])
    # No matching shift exists yet, so this row will resolve as an error —
    # this test only verifies the end-to-end wiring (upload -> list -> get -> confirm),
    # not the resolution rules (covered in test_import_sessions_service.py).
    resp = _upload(client, _token(admin), _to_bytes(wb))
    assert resp.status_code == 200
    session_id = resp.json()["session_id"]
    assert resp.json()["preview"]["assignments"][0]["action"] == "error"

    list_resp = client.get(
        "/api/import/sessions", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    summary = next(s for s in list_resp.json() if s["id"] == session_id)
    assert summary["row_summary"]["assignments"] == 1

    detail_resp = client.get(
        f"/api/import/sessions/{session_id}", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    assert detail_resp.json()["parsed_state"]["assignments"][0]["action"] == "error"

    confirm_resp = client.post(
        f"/api/import/sessions/{session_id}/confirm",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert confirm_resp.status_code == 200
    assert confirm_resp.json()["created"] == 0  # error row, nothing created
    assert admin_session.execute(select(DutyAssignment)).scalars().all() == []
