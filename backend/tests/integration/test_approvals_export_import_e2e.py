from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from datetime import date as date_type

import openpyxl
from sqlalchemy import select

import app.services.import_parsers.v1_standard  # noqa: F401 -- registers the v1_standard parser
from app.db.models import (
    DutyAssignment,
    DutyLocation,
    DutyManagerScope,
    DutyType,
    ExemptionRequest,
    ExemptionType,
    PersonalConstraint,
    RangeExcusalRequest,
    Soldier,
    SoldierEnrollmentRequest,
    SoldierExemption,
    SoldierFieldUpdate,
    SwapCandidate,
    SwapManagerApproval,
    SwapRequest,
)
from app.services.hierarchy import create_node
from app.services.import_sessions import confirm_session, create_session
from tests.helpers import auth_headers, create_range_event, create_range_location, create_soldier


def _uid() -> str:
    return uuid.uuid4().hex[:8]


def _token(soldier: Soldier) -> str:
    return auth_headers(soldier)["Authorization"].split(" ", 1)[1]


def _admin_headers(soldier: Soldier) -> dict[str, str]:
    return {"Authorization": f"Bearer {_token(soldier)}"}


def _make_assignment(session, *, soldier: Soldier, node=None, start_date=date_type(2026, 4, 1)) -> DutyAssignment:
    dt = DutyType(name=f"dt_{_uid()}", score_per_day=1)
    loc = DutyLocation(name=f"loc_{_uid()}")
    session.add_all([dt, loc])
    session.flush()
    assignment = DutyAssignment(
        duty_type_id=dt.id,
        duty_location_id=loc.id,
        soldier_id=soldier.id,
        start_date=start_date,
        end_date=start_date,
        status="published",
    )
    session.add(assignment)
    session.flush()
    return assignment


def _export(client, admin: Soldier, sheet: str):
    resp = client.get(
        f"/api/approvals/export?sheets={sheet}", headers=_admin_headers(admin)
    )
    assert resp.status_code == 200
    return resp.content


def test_personal_constraint_export_import_round_trip(admin_session, client):
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}", hierarchy_node_id=node.id)
    decider = create_soldier(admin_session, personal_number=f"dec_{_uid()}")
    original = PersonalConstraint(
        soldier_id=soldier.id, start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        reason="round trip test", status="approved", decided_by=decider.id,
        decided_at=datetime.now(UTC), decision_note="ok",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    # Export applies the same reason-visibility redaction as the interactive
    # constraint endpoints (can_see_private): give the admin actor
    # duty-manager scope over the soldier's node so the reason round-trips
    # instead of being (correctly) redacted to None and then blanking the
    # existing DB value on re-import.
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    admin_session.commit()
    xlsx_bytes = _export(client, admin, "personal_constraints")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["personal_constraints"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(original.id)

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.status == "approved"
    assert original.decided_by == decider.id
    assert original.decision_note == "ok"
    assert original.start_date == date_type(2024, 6, 15)
    assert original.end_date == date_type(2024, 6, 16)
    assert original.reason == "round trip test"


def test_personal_constraint_import_accepts_new_split_approval_statuses(admin_session, client):
    # Regression test: the two-step approval split (Task 9) added
    # "pending_commander" / "pending_duty_manager" as PersonalConstraint.status
    # values, replacing the old single "pending" state. The importer's status
    # validation in resolve_personal_constraints previously still checked
    # against the old 3-state set ("pending", "approved", "rejected"), so a
    # row carrying "pending_commander" would be silently marked action="error"
    # instead of being accepted for update. This must no longer happen.
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}", hierarchy_node_id=node.id)
    original = PersonalConstraint(
        soldier_id=soldier.id, start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        reason="pending commander approval", status="pending_commander",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    admin_session.commit()
    xlsx_bytes = _export(client, admin, "personal_constraints")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["personal_constraints"][0]
    assert row["errors"] == []
    assert row["action"] == "update"
    assert row["status"] == "pending_commander"

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)
    assert original.status == "pending_commander"

    # No duplicate row was created by the round trip.
    all_rows = admin_session.execute(select(PersonalConstraint)).scalars().all()
    assert len(all_rows) == 1


def test_personal_constraint_import_coerces_legacy_pending_status(admin_session, client):
    # Regression test: approvals workbooks exported before the two-step
    # approval split (Task 9) still carry the literal legacy value "pending"
    # for personal_constraints.status. resolve_personal_constraints must
    # coerce this legacy value to "pending_commander" before validating it,
    # rather than hard-rejecting the row with "סטטוס לא תקין 'pending'".
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}", hierarchy_node_id=node.id)
    original = PersonalConstraint(
        soldier_id=soldier.id, start_date=date_type(2024, 6, 15), end_date=date_type(2024, 6, 16),
        reason="legacy pending export", status="pending_commander",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    admin_session.commit()
    xlsx_bytes = _export(client, admin, "personal_constraints")

    # Simulate a pre-Task-9 export: rewrite the status cell to the legacy
    # literal "pending" value that older exports would still carry.
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["personal_constraints"]
    header = [cell.value for cell in ws[1]]
    status_col = header.index("status") + 1
    ws.cell(row=2, column=status_col, value="pending")
    buf = io.BytesIO()
    wb.save(buf)
    legacy_bytes = buf.getvalue()

    sess = create_session(admin_session, filename="legacy.xlsx", content=legacy_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["personal_constraints"][0]
    assert row["errors"] == []
    assert row["action"] == "update"
    assert row["status"] == "pending_commander"

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)
    assert original.status == "pending_commander"


def test_soldier_field_update_export_import_round_trip(admin_session, client):
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}")
    decider = create_soldier(admin_session, personal_number=f"dec_{_uid()}")
    original = SoldierFieldUpdate(
        soldier_id=soldier.id, field_name="phone", new_value="0501234567",
        previous_value="0500000000", status="approved", decided_by=decider.id,
        decided_at=datetime.now(UTC), decision_note="ok",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    xlsx_bytes = _export(client, admin, "soldier_field_updates")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["soldier_field_updates"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(original.id)

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.status == "approved"
    assert original.decided_by == decider.id
    assert original.decision_note == "ok"
    assert original.field_name == "phone"
    assert original.new_value == "0501234567"
    assert original.previous_value == "0500000000"

    all_rows = admin_session.execute(select(SoldierFieldUpdate)).scalars().all()
    assert len(all_rows) == 1


def test_soldier_enrollment_request_export_import_round_trip(admin_session, client):
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}")
    decider = create_soldier(admin_session, personal_number=f"dec_{_uid()}")
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    original = SoldierEnrollmentRequest(
        soldier_id=soldier.id, requested_node_id=node.id, status="approved",
        decided_by=decider.id, decided_at=datetime.now(UTC), decision_note="ok",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    xlsx_bytes = _export(client, admin, "soldier_enrollment_requests")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["soldier_enrollment_requests"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(original.id)

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.status == "approved"
    assert original.decided_by == decider.id
    assert original.decision_note == "ok"
    assert original.requested_node_id == node.id

    all_rows = admin_session.execute(select(SoldierEnrollmentRequest)).scalars().all()
    assert len(all_rows) == 1


def test_soldier_exemption_export_import_round_trip(admin_session, client):
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}")
    granter = create_soldier(admin_session, personal_number=f"gr_{_uid()}")
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=False, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    original = SoldierExemption(
        soldier_id=soldier.id, exemption_type_id=et.id, start_date=date_type(2026, 2, 1),
        end_date=date_type(2026, 2, 10), reason="פציעה", granted_by=granter.id,
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    xlsx_bytes = _export(client, admin, "soldier_exemptions")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["soldier_exemptions"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(original.id)

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.reason == "פציעה"
    assert original.granted_by == granter.id
    assert original.start_date == date_type(2026, 2, 1)
    assert original.end_date == date_type(2026, 2, 10)
    assert original.revoked_at is None

    all_rows = admin_session.execute(select(SoldierExemption)).scalars().all()
    assert len(all_rows) == 1


def test_exemption_request_export_import_round_trip(admin_session, client):
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}", hierarchy_node_id=node.id)
    commander = create_soldier(admin_session, personal_number=f"cmd_{_uid()}")
    decider = create_soldier(admin_session, personal_number=f"dec_{_uid()}")
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=True, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    original = ExemptionRequest(
        soldier_id=soldier.id, exemption_type_id=et.id, start_date=date_type(2026, 3, 1),
        end_date=date_type(2026, 3, 5), reason="בדיקה רפואית", status="approved",
        commander_approved_by=commander.id, decided_by=decider.id, decision_note="ok",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    # Same redaction policy as constraints (can_see_private) — grant scope so
    # the reason round-trips instead of being blanked on re-import.
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    admin_session.commit()
    xlsx_bytes = _export(client, admin, "exemption_requests")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["exemption_requests"][0]
    assert row["action"] == "update"
    assert row["existing_id"] == str(original.id)

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.status == "approved"
    assert original.commander_approved_by == commander.id
    assert original.decided_by == decider.id
    assert original.decision_note == "ok"
    assert original.reason == "בדיקה רפואית"

    all_rows = admin_session.execute(select(ExemptionRequest)).scalars().all()
    assert len(all_rows) == 1


def test_swap_request_export_import_round_trip_preserves_approval_log(admin_session, client):
    node = create_node(admin_session, level="unit", name=f"n_{_uid()}", parent_id=None)
    requester = create_soldier(admin_session, personal_number=f"r_{_uid()}", hierarchy_node_id=node.id)
    covering = create_soldier(admin_session, personal_number=f"c_{_uid()}", hierarchy_node_id=node.id)
    commander = create_soldier(admin_session, personal_number=f"cm_{_uid()}")
    node.commander_id = commander.id
    admin_session.flush()
    assignment = _make_assignment(admin_session, soldier=requester, node=node)
    # "pending_approval" no longer exists as a SwapRequest.status — a swap
    # mid-approval is now "open" with a live SwapCandidate carrying its own
    # soldier_side_approved.
    original = SwapRequest(
        duty_assignment_id=assignment.id, duty_date=assignment.start_date,
        requesting_soldier_id=requester.id,
        status="open", requester_side_approved=True,
    )
    admin_session.add(original)
    admin_session.flush()
    candidate = SwapCandidate(
        swap_request_id=original.id, soldier_id=covering.id, source="invited",
        status="accepted", soldier_side_approved=True,
    )
    admin_session.add(candidate)
    admin_session.flush()
    decision = SwapManagerApproval(
        swap_request_id=original.id, side="requester", commander_id=commander.id,
        approver_kind="commander", approved=True, approved_by=commander.id, approved_at=datetime.now(UTC),
    )
    admin_session.add(decision)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    xlsx_bytes = _export(client, admin, "swap_requests")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    row = sess.parsed_state["swap_requests"][0]
    assert row["action"] == "update"
    assert len(row["approval_log"]) == 1
    assert row["approval_log"][0]["side"] == "requester"
    assert row["approval_log"][0]["outcome"] == "approved"

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()

    rows = admin_session.execute(
        select(SwapManagerApproval).where(SwapManagerApproval.swap_request_id == original.id)
    ).scalars().all()
    assert len(rows) == 1  # idempotent -- re-confirming didn't duplicate the decision row
    assert rows[0].approved is True
    assert rows[0].commander_id == commander.id
    # Note carried forward from an earlier review: the swap_requests export's
    # approval_log column logs the decision-log row's commander_id (the actor who
    # made the decision) rather than a separately-named approved_by/rejected_by
    # field. commander_id and approved_by are always the same value for any given
    # decision row under the current codebase's design, so both survive the round
    # trip identically.
    assert rows[0].approved_by == commander.id

    admin_session.refresh(original)
    assert original.status == "open"
    assert original.requester_side_approved is True

    admin_session.refresh(candidate)
    assert candidate.soldier_side_approved is True
    assert candidate.soldier_id == covering.id
    assert candidate.status == "accepted"


def test_range_excusal_request_export_import_round_trip(admin_session, client):
    """Round-trip test for approvals_export's range_excusal_requests sheet
    (Task 10). Uses an *approved* excusal whose linked RangeAssignment has
    already been deleted (matching what range_excusal.py's approval flow
    actually does, and what a real approvals export of already-approved data
    looks like) — this is the exact scenario Finding 3 fixed: without that
    fix, re-importing this row would error out on "soldier not found"
    (soldier_personal_number is necessarily blank once the assignment is
    gone) instead of resolving as a clean update."""
    node = create_node(admin_session, level="group", name=f"n_{_uid()}", parent_id=None)
    loc = create_range_location(admin_session, name=f"loc_{_uid()}")
    soldier = create_soldier(admin_session, personal_number=f"s_{_uid()}", hierarchy_node_id=node.id)
    decider = create_soldier(admin_session, personal_number=f"dec_{_uid()}")
    event = create_range_event(
        admin_session, hierarchy_node=node, range_location=loc, event_date=date_type(2024, 6, 20),
    )
    original = RangeExcusalRequest(
        range_assignment_id=None,  # deleted by the approval flow, per Finding 3
        range_event_id=event.id,
        requested_by=soldier.id,
        reason="פציעה",
        status="approved",
        decided_by=decider.id,
        decided_at=datetime.now(UTC),
        decision_note="אושר",
    )
    admin_session.add(original)
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    xlsx_bytes = _export(client, admin, "range_excusal_requests")

    sess = create_session(admin_session, filename="roundtrip.xlsx", content=xlsx_bytes, actor=admin, parser_id="v1_standard")
    # The export contains every range_excusal_request in the DB (not just this
    # test's), so match by existing_id rather than assuming index 0 — other
    # tests sharing this worker's DB may have already created their own rows.
    row = next(r for r in sess.parsed_state["range_excusal_requests"] if r["existing_id"] == str(original.id))
    assert row["errors"] == []
    assert row["action"] == "update"

    confirm_session(admin_session, session_id=sess.id, actor=admin)
    admin_session.commit()
    admin_session.refresh(original)

    assert original.status.value == "approved"
    assert original.decided_by == decider.id
    assert original.decision_note == "אושר"

    # No duplicate of this specific row was created by the round trip.
    matches = admin_session.execute(
        select(RangeExcusalRequest).where(RangeExcusalRequest.id == original.id)
    ).scalars().all()
    assert len(matches) == 1
