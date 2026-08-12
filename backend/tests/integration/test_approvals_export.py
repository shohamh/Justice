from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta, timezone

import openpyxl

from app.db.models import (
    DutyAssignment,
    DutyLocation,
    DutyManagerScope,
    DutyType,
    ExemptionRequest,
    ExemptionType,
    PersonalConstraint,
    SoldierEnrollmentRequest,
    SoldierExemption,
    SoldierFieldUpdate,
    SwapCandidate,
    SwapManagerApproval,
    SwapRequest,
)
from app.services.hierarchy import create_node as create_hierarchy_node
from tests.helpers import auth_headers, create_soldier


def _uid() -> str:
    return uuid.uuid4().hex[:8]


def _token(soldier) -> str:
    return auth_headers(soldier)["Authorization"].split(" ", 1)[1]


def test_export_personal_constraints_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    soldier = create_soldier(
        admin_session, personal_number=f"sld_{_uid()}", hierarchy_node_id=node.id
    )
    # Give the admin actor duty-manager scope over the soldier's node so
    # can_see_private grants visibility — the export writer now applies the
    # same privacy redaction as the interactive read endpoints, and a bare
    # admin role is deliberately NOT a blanket bypass for private fields.
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    constraint = PersonalConstraint(
        soldier_id=soldier.id,
        start_date=date(2026, 1, 1),
        end_date=date(2026, 1, 5),
        reason="חופשה",
        status="approved",
        decided_by=admin.id,
        decision_note="ok",
    )
    admin_session.add(constraint)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=personal_constraints",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["personal_constraints"]
    ws = wb["personal_constraints"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "id", "soldier_personal_number", "soldier_name", "start_date", "end_date",
        "reason", "status", "decided_by_personal_number", "decision_note", "created_at",
    ]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(constraint.id))
    assert row[1] == soldier.personal_number
    assert row[2] == soldier.full_name
    assert row[3] == "2026-01-01"
    assert row[4] == "2026-01-05"
    assert row[5] == "חופשה"
    assert row[6] == "approved"
    assert row[7] == admin.personal_number
    assert row[8] == "ok"


def test_export_personal_constraints_sheet_redacts_reason_when_actor_out_of_scope(client, admin_session):
    """An admin export actor with no commander/duty-manager scope over the
    soldier's node must NOT see the private `reason` text — matching the
    same `can_see_private` policy the interactive constraints endpoints
    enforce (see app/routes/constraints.py). Admin role alone is not a
    blanket bypass for this field."""
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    soldier = create_soldier(
        admin_session, personal_number=f"sld_{_uid()}", hierarchy_node_id=node.id
    )
    constraint = PersonalConstraint(
        soldier_id=soldier.id,
        start_date=date(2026, 1, 1),
        end_date=date(2026, 1, 5),
        reason="חופשה",
        status="approved",
    )
    admin_session.add(constraint)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=personal_constraints",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["personal_constraints"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(constraint.id))
    assert row[5] is None


def test_export_swap_requests_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    requester = create_soldier(admin_session, personal_number=f"req_{_uid()}")
    covering = create_soldier(admin_session, personal_number=f"cov_{_uid()}")
    commander = create_soldier(admin_session, personal_number=f"cmd_{_uid()}", role="commander")
    rejecting_commander = create_soldier(admin_session, personal_number=f"rcmd_{_uid()}", role="commander")

    dt = DutyType(name=f"dt_{_uid()}", score_per_day=1)
    loc = DutyLocation(name=f"loc_{_uid()}")
    admin_session.add_all([dt, loc])
    admin_session.flush()
    assignment = DutyAssignment(
        duty_type_id=dt.id, duty_location_id=loc.id, soldier_id=requester.id,
        start_date=date(2026, 4, 1), end_date=date(2026, 4, 1), status="published",
    )
    admin_session.add(assignment)
    admin_session.flush()

    # "pending_approval" no longer exists as a SwapRequest.status — a swap
    # mid-approval is now "open" with a live SwapCandidate carrying its own
    # soldier_side_approved. covering_side_approved=False here means the
    # candidate has already engaged (status="accepted", not just an
    # unanswered invite) but hasn't confirmed their own side yet.
    swap = SwapRequest(
        duty_assignment_id=assignment.id, duty_date=assignment.start_date,
        requesting_soldier_id=requester.id,
        status="open", reason="סיבה אישית",
        requester_side_approved=True,
        decision_note="ממתין לאישור צד שני",
    )
    admin_session.add(swap)
    admin_session.flush()

    candidate = SwapCandidate(
        swap_request_id=swap.id, soldier_id=covering.id, source="invited",
        status="accepted", soldier_side_approved=False,
    )
    admin_session.add(candidate)
    admin_session.flush()

    approved_at = datetime(2026, 4, 1, 8, 30, tzinfo=timezone.utc)
    approval = SwapManagerApproval(
        swap_request_id=swap.id, side="requester", commander_id=commander.id,
        approved=True, approved_by=commander.id, approved_at=approved_at,
        approver_kind="commander",
    )
    admin_session.add(approval)

    rejected_at = approved_at + timedelta(minutes=5)
    rejection = SwapManagerApproval(
        swap_request_id=swap.id, swap_candidate_id=candidate.id, side="covering",
        commander_id=rejecting_commander.id,
        rejected=True, rejected_by=rejecting_commander.id, rejected_at=rejected_at,
        approver_kind="commander",
    )
    admin_session.add(rejection)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=swap_requests",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["swap_requests"]
    ws = wb["swap_requests"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "id", "requesting_personal_number", "requesting_name", "target_personal_number",
        "covering_personal_number", "duty_date", "status", "reason",
        "requester_side_approved", "covering_side_approved",
        "rejected_by_personal_number", "decision_note", "approval_log", "created_at", "updated_at",
    ]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(swap.id))
    assert row[1] == requester.personal_number
    assert row[2] == requester.full_name
    assert row[3] is None
    assert row[4] == covering.personal_number
    assert row[5] == "2026-04-01"
    assert row[6] == "open"
    assert row[7] == "סיבה אישית"
    assert row[8] is True
    assert row[9] is False
    assert row[10] is None
    assert row[11] == "ממתין לאישור צד שני"

    approval_log = row[12]
    segments = approval_log.split(";")
    assert len(segments) == 2
    approved_segment = f"requester:commander:{commander.personal_number}:approved:{approved_at.isoformat()}"
    rejected_segment = f"covering:commander:{rejecting_commander.personal_number}:rejected:{rejected_at.isoformat()}"
    assert approved_segment in segments
    assert rejected_segment in segments


def test_export_defaults_to_all_six_sheets(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    resp = client.get(
        "/api/approvals/export", headers={"Authorization": f"Bearer {_token(admin)}"}
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {
        "swap_requests", "exemption_requests", "soldier_field_updates",
        "soldier_enrollment_requests", "personal_constraints", "soldier_exemptions",
    }


def test_export_soldier_field_updates_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    soldier = create_soldier(admin_session, personal_number=f"sld_{_uid()}")
    update = SoldierFieldUpdate(
        soldier_id=soldier.id,
        field_name="phone",
        new_value="0501234567",
        previous_value="0500000000",
        status="pending",
    )
    admin_session.add(update)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=soldier_field_updates",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["soldier_field_updates"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(update.id))
    assert row[1] == soldier.personal_number
    assert row[3] == "phone"
    assert row[4] == "0501234567"
    assert row[5] == "0500000000"
    assert row[6] == "pending"


def test_export_soldier_enrollment_requests_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    soldier = create_soldier(admin_session, personal_number=f"sld_{_uid()}")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    req = SoldierEnrollmentRequest(soldier_id=soldier.id, requested_node_id=node.id, status="pending")
    admin_session.add(req)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=soldier_enrollment_requests",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["soldier_enrollment_requests"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(req.id))
    assert row[1] == soldier.personal_number
    assert row[3] == node.name
    assert row[4] == "pending"


def test_export_soldier_exemptions_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    soldier = create_soldier(admin_session, personal_number=f"sld_{_uid()}")
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=False, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    exemption = SoldierExemption(
        soldier_id=soldier.id,
        exemption_type_id=et.id,
        start_date=date(2026, 2, 1),
        reason="פציעה",
        granted_by=admin.id,
    )
    admin_session.add(exemption)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=soldier_exemptions",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["soldier_exemptions"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(exemption.id))
    assert row[1] == soldier.personal_number
    assert row[3] == et.name
    assert row[4] == "2026-02-01"
    assert row[6] == "פציעה"
    assert row[7] == admin.personal_number


def test_export_exemption_requests_sheet(client, admin_session):
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    soldier = create_soldier(
        admin_session, personal_number=f"sld_{_uid()}", hierarchy_node_id=node.id
    )
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=False, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    req = ExemptionRequest(
        soldier_id=soldier.id,
        exemption_type_id=et.id,
        start_date=date(2026, 3, 1),
        reason="בדיקה רפואית",
        status="pending",
    )
    admin_session.add(req)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=exemption_requests",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["exemption_requests"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(req.id))
    assert row[1] == soldier.personal_number
    assert row[3] == et.name
    assert row[6] == "בדיקה רפואית"
    assert row[7] == "pending"


def test_export_exemption_requests_sheet_survives_permanent_request(client, admin_session):
    """Regression test: a permanent exemption request (start_date=None) used
    to crash the whole XLSX export (r.start_date.isoformat() with no None
    guard) for admins as soon as any such row existed, not just the row
    itself failing to render."""
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    soldier = create_soldier(
        admin_session, personal_number=f"sld_{_uid()}", hierarchy_node_id=node.id
    )
    admin_session.add(DutyManagerScope(duty_manager_id=admin.id, hierarchy_node_id=node.id))
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=False, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    req = ExemptionRequest(
        soldier_id=soldier.id,
        exemption_type_id=et.id,
        start_date=None,
        reason="פטור קבוע",
        status="pending",
    )
    admin_session.add(req)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=exemption_requests",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["exemption_requests"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(req.id))
    assert not row[4]


def test_export_exemption_requests_sheet_redacts_reason_when_actor_out_of_scope(client, admin_session):
    """Same privacy policy as constraints: an export actor with no
    commander/duty-manager scope over the soldier's node must not see the
    private `reason` text (mirrors exemption_requests.py's include_sensitive
    check, which is also driven by can_see_private)."""
    admin = create_soldier(admin_session, personal_number=f"adm_{_uid()}", role="admin")
    node = create_hierarchy_node(admin_session, level="group", name=f"מדור_{_uid()}", parent_id=None)
    soldier = create_soldier(
        admin_session, personal_number=f"sld_{_uid()}", hierarchy_node_id=node.id
    )
    et = ExemptionType(name=f"et_{_uid()}", is_global=False, is_medical=False, is_commander_exemption=False)
    admin_session.add(et)
    admin_session.flush()
    req = ExemptionRequest(
        soldier_id=soldier.id,
        exemption_type_id=et.id,
        start_date=date(2026, 3, 1),
        reason="בדיקה רפואית",
        status="pending",
    )
    admin_session.add(req)
    admin_session.commit()

    resp = client.get(
        "/api/approvals/export?sheets=exemption_requests",
        headers={"Authorization": f"Bearer {_token(admin)}"},
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    rows = list(wb["exemption_requests"].iter_rows(min_row=2, values_only=True))
    row = next(r for r in rows if r[0] == str(req.id))
    assert row[1] == soldier.personal_number
    assert row[6] is None
