from __future__ import annotations

import io
import uuid
from decimal import Decimal

import openpyxl
import pytest

from app.db.models import DutyLocation, DutyType
from app.services.duty_config import create_duty_type
from tests.helpers import auth_headers, create_node, create_soldier


def make_xlsx_bytes(soldiers=None, assignments=None) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if soldiers:
        ws = wb.create_sheet("soldiers")
        ws.append(["personal_number", "full_name", "rank"])
        for row in soldiers:
            ws.append(row)
    if assignments:
        ws = wb.create_sheet("assignments")
        ws.append(["personal_number", "duty_type_name", "start_date", "end_date", "is_reserve"])
        for row in assignments:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, xlsx: bytes, token: str):
    return client.post(
        "/api/import/preview",
        files={"file": ("import.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )


def test_preview_new_soldier(client, admin_session):
    node = create_node(admin_session, level="branch", name="ie_node_001")
    dm = create_soldier(admin_session, personal_number="ie_dm_001", role="duty_manager", hierarchy_node_id=node.id)
    xlsx = make_xlsx_bytes(soldiers=[["ie_new_001", "ישראל ישראלי", "רב"]])
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]
    resp = _upload(client, xlsx, token)
    assert resp.status_code == 200
    soldiers = resp.json()["soldiers"]
    assert len(soldiers) == 1
    assert soldiers[0]["action"] == "new"
    assert soldiers[0]["personal_number"] == "ie_new_001"


def test_preview_duplicate_soldier_is_update(client, admin_session):
    node = create_node(admin_session, level="branch", name="ie_node_002")
    dm = create_soldier(admin_session, personal_number="ie_dm_002", role="duty_manager", hierarchy_node_id=node.id)
    existing = create_soldier(admin_session, personal_number="ie_existing_002", hierarchy_node_id=node.id)
    xlsx = make_xlsx_bytes(soldiers=[[existing.personal_number, "שם חדש", None]])
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]
    resp = _upload(client, xlsx, token)
    assert resp.json()["soldiers"][0]["action"] == "update"


def test_apply_creates_soldier(client, admin_session):
    node = create_node(admin_session, level="branch", name="ie_node_003")
    dm = create_soldier(admin_session, personal_number="ie_dm_003", role="duty_manager", hierarchy_node_id=node.id)
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]
    resp = client.post(
        "/api/import/apply",
        json={
            "soldiers": [{
                "row": 2, "action": "new",
                "personal_number": "ie_apply_003", "full_name": "טסט יחידה",
                "rank": None, "gender": None, "is_officer": None,
                "hierarchy_node_id": str(node.id), "enrolled_at": None,
                "enlistment_date": None, "phone": None, "email": None, "existing_id": None,
            }],
            "assignments": [],
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 1
    assert resp.json()["errors"] == []


def test_apply_rejects_out_of_scope_hierarchy_node(client, admin_session):
    """A duty manager scoped to unit A must not be able to import a soldier
    into unit B via /import/apply."""
    node_a = create_node(admin_session, level="branch", name="ie_node_scope_a")
    node_b = create_node(admin_session, level="branch", name="ie_node_scope_b")
    dm = create_soldier(admin_session, personal_number="ie_dm_scope", role="duty_manager", hierarchy_node_id=node_a.id)
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]

    resp = client.post(
        "/api/import/apply",
        json={
            "soldiers": [{
                "row": 2, "action": "new",
                "personal_number": "ie_apply_scope", "full_name": "טסט חריגה",
                "rank": None, "gender": None, "is_officer": None,
                "hierarchy_node_id": str(node_b.id), "enrolled_at": None,
                "enlistment_date": None, "phone": None, "email": None, "existing_id": None,
            }],
            "assignments": [],
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 403
    assert "out_of_scope_rows" in resp.json()["detail"]


def test_apply_notifies_soldier_of_new_assignment(client, admin_session):
    from app.db.models import Notification, NotificationType

    node = create_node(admin_session, level="branch", name="ie_node_notif")
    dm = create_soldier(admin_session, personal_number="ie_dm_notif", role="duty_manager", hierarchy_node_id=node.id)
    soldier = create_soldier(admin_session, personal_number="ie_soldier_notif", hierarchy_node_id=node.id)
    dt = create_duty_type(admin_session, name=f"dt_notif_{uuid.uuid4().hex[:8]}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_notif_{uuid.uuid4().hex[:8]}")
    admin_session.add(loc)
    admin_session.commit()
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]

    resp = client.post(
        "/api/import/apply",
        json={
            "soldiers": [],
            "assignments": [{
                "row": 2, "action": "new",
                "resolved_soldier_id": str(soldier.id),
                "resolved_duty_type_id": str(dt.id),
                "start_date": "2024-06-15", "end_date": "2024-06-16",
                "is_reserve": False,
            }],
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 1

    admin_session.expire_all()
    notif = admin_session.query(Notification).filter_by(
        soldier_id=soldier.id, type=NotificationType.assignment_created,
    ).one_or_none()
    assert notif is not None


def test_apply_succeeds_even_if_notification_fails(client, admin_session, monkeypatch):
    """If sending the post-import notification raises, the import's own side
    effects (soldiers/assignments/audit row) already committed successfully,
    so the endpoint must still return 200 with the correct counts rather than
    an unhandled 500."""
    import app.routes.import_excel as import_excel_module

    def _boom(*args, **kwargs):
        raise RuntimeError("simulated notification failure")

    monkeypatch.setattr(import_excel_module, "create_notification", _boom)

    node = create_node(admin_session, level="branch", name="ie_node_notif_fail")
    dm = create_soldier(admin_session, personal_number="ie_dm_notif_fail", role="duty_manager", hierarchy_node_id=node.id)
    soldier = create_soldier(admin_session, personal_number="ie_soldier_notif_fail", hierarchy_node_id=node.id)
    dt = create_duty_type(admin_session, name=f"dt_notif_fail_{uuid.uuid4().hex[:8]}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_notif_fail_{uuid.uuid4().hex[:8]}")
    admin_session.add(loc)
    admin_session.commit()
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]

    resp = client.post(
        "/api/import/apply",
        json={
            "soldiers": [],
            "assignments": [{
                "row": 2, "action": "new",
                "resolved_soldier_id": str(soldier.id),
                "resolved_duty_type_id": str(dt.id),
                "start_date": "2024-06-15", "end_date": "2024-06-16",
                "is_reserve": False,
            }],
        },
        headers={"Authorization": f"Bearer {token}"},
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 1

    # The assignment itself was durably committed despite the notification failure.
    from app.db.models import DutyAssignment
    admin_session.expire_all()
    assignment = admin_session.query(DutyAssignment).filter_by(soldier_id=soldier.id).one_or_none()
    assert assignment is not None


def test_template_download(client, admin_session):
    node = create_node(admin_session, level="branch", name="ie_node_004")
    dm = create_soldier(admin_session, personal_number="ie_dm_004", role="duty_manager", hierarchy_node_id=node.id)
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/template", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == {
        "soldiers", "duty_shifts", "assignments",
        "duty_locations", "hierarchy", "duty_types", "exemption_types", "shift_templates",
    }
    headers = [c.value for c in next(wb["assignments"].iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "personal_number", "full_name", "duty_type_name", "duty_location_name",
        "start_date", "end_date", "start_time", "end_time", "is_reserve", "notes",
    ]


def test_template_download_includes_shift_templates_sheet(client, admin_session):
    node = create_node(admin_session, level="branch", name="ie_node_005")
    dm = create_soldier(admin_session, personal_number="ie_dm_005", role="duty_manager", hierarchy_node_id=node.id)
    token = auth_headers(dm)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/template", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "shift_templates" in wb.sheetnames
    headers = [c.value for c in next(wb["shift_templates"].iter_rows(min_row=1, max_row=1))]
    assert headers == [
        "name", "duty_type_name", "duty_location_name", "recurrence_type", "weekdays",
        "start_time", "end_time", "required_count", "auto_roll", "auto_roll_until",
        "duration_days", "notes", "eligible_units",
    ]


def test_export_current_data_includes_shift_templates(client, admin_session):
    from app.services.shift_templates import create_template
    from app.db.models import DutyLocation

    dt = create_duty_type(admin_session, name=f"dt_export_{uuid.uuid4().hex[:8]}", score_per_day=Decimal("1.00"))
    loc = DutyLocation(name=f"loc_export_{uuid.uuid4().hex[:8]}")
    admin_session.add(loc)
    admin_session.flush()
    tpl_name = f"tpl_export_{uuid.uuid4().hex[:8]}"
    create_template(
        admin_session, name=tpl_name, duty_type_id=dt.id, duty_location_id=loc.id,
        recurrence_type="weekdays", weekdays=[], required_count=1,
    )
    admin_session.commit()

    admin = create_soldier(admin_session, personal_number=f"adm_export_{uuid.uuid4().hex[:8]}", role="admin")
    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/export", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "shift_templates" in wb.sheetnames
    rows = list(wb["shift_templates"].iter_rows(min_row=2, values_only=True))
    names = [r[0] for r in rows]
    assert tpl_name in names
