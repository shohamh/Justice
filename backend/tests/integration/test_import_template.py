from __future__ import annotations

import io

import openpyxl

from tests.helpers import auth_headers, create_soldier


def test_template_includes_all_six_sheets(client, admin_session):
    admin = create_soldier(admin_session, personal_number="tmpl-admin", role="admin")
    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/template", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) >= {
        "חיילים", "משמרות", "מיקומי תורנויות", "היררכיה", "סוגי תפקידים", "סוגי פטורים",
    }


def test_template_includes_range_sheets(client, admin_session):
    admin = create_soldier(admin_session, personal_number="tmpl-admin-range", role="admin")
    token = auth_headers(admin)["Authorization"].split(" ", 1)[1]
    resp = client.get("/api/import/template", headers={"Authorization": f"Bearer {token}"})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    for name in (
        "מיקומי מטווח", "ימי מטווח", "שיבוצי מטווח",
        "כשירויות מטווח", "בקשות היעדרות",
    ):
        assert name in wb.sheetnames
